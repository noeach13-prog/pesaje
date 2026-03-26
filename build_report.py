import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.load(open(r'C:\Users\EliteBook\Pesaje\report_data.json', encoding='utf-8'))
wb = Workbook()

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2F5496')
num_font = Font(name='Arial', size=10)
adj_font = Font(name='Arial', size=10, color='C00000', bold=True)
total_fill = PatternFill('solid', fgColor='D6E4F0')
total_font = Font(name='Arial', bold=True, size=11)
neg_font = Font(name='Arial', size=10, color='C00000')
thin = Side(style='thin', color='B0B0B0')
border = Border(bottom=thin)

def style_header(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(c)].width = w

# ── HOJA 1: TOTALES POR DIA ──
ws = wb.active
ws.title = 'Totales por Dia'
style_header(ws,
    ['Dia', 'Stock Engine (g)', 'Ajuste MT (g)', 'Stock Corregido (g)',
     'VDP (g)', 'Latas', 'Lid Discount (g)', 'Total Dia (g)', 'Correcciones'],
    [6, 16, 14, 18, 10, 7, 16, 16, 12])

for r, d in enumerate(data['day_results'], 2):
    vals = [d['day'], d['stock_engine'], d['adjustment'], d['stock_corrected'],
            d['vdp'], d['latas'], d['lid'], d['total_corrected'], d['n_corrections']]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = num_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if c in (2, 3, 4, 5, 7, 8):
            cell.number_format = '#,##0'
        if c == 3:
            cell.number_format = '+#,##0;-#,##0;"-"'
            if v != 0: cell.font = adj_font
        if c in (2, 4, 8) and isinstance(v, (int, float)) and v < 0:
            cell.font = neg_font

tr = len(data['day_results']) + 2
ws.cell(row=tr, column=1, value='TOTAL').font = total_font
for col in range(1, 10):
    ws.cell(row=tr, column=col).fill = total_fill
    ws.cell(row=tr, column=col).font = total_font
    ws.cell(row=tr, column=col).alignment = Alignment(horizontal='center')
for ci, cl in [(2,'B'), (3,'C'), (4,'D'), (5,'E'), (6,'F'), (7,'G'), (8,'H')]:
    ws.cell(row=tr, column=ci).value = f'=SUM({cl}2:{cl}{tr-1})'
    ws.cell(row=tr, column=ci).number_format = '#,##0'

# ── HOJA 2: CORRECCIONES MULTI-TURNO ──
ws2 = wb.create_sheet('Correcciones Multi-Turno')
style_header(ws2,
    ['Dia', 'Sabor', 'Valor Engine (g)', 'Valor Corregido (g)', 'Delta (g)', 'Tipo'],
    [6, 22, 17, 19, 14, 45])

row = 2
for dn in sorted(data['mt_corrections'].keys(), key=lambda x: int(x)):
    for (fname, ov, nv, reason) in data['mt_corrections'][dn]:
        vals = [int(dn), fname, ov, nv, nv - ov, reason]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.font = num_font if c < 6 else Font(name='Arial', size=9)
            cell.border = border
            if c in (3, 4): cell.number_format = '#,##0'
            if c == 5:
                cell.number_format = '+#,##0;-#,##0;"-"'
                cell.font = adj_font
        row += 1

# ── HOJA 3: LATAS ABIERTAS ──
ws3 = wb.create_sheet('Latas Abiertas')
style_header(ws3,
    ['Dia', 'Cantidad', 'Lid Discount (g)', 'Detalle'],
    [6, 10, 16, 70])

row = 2
for d in data['day_results']:
    if d['latas'] > 0:
        ws3.cell(row=row, column=1, value=d['day']).font = num_font
        ws3.cell(row=row, column=2, value=d['latas']).font = num_font
        c3 = ws3.cell(row=row, column=3, value=d['lid'])
        c3.font = num_font; c3.number_format = '#,##0'
        detail = '; '.join(d['opening_events'][:8]) if d['opening_events'] else ''
        ws3.cell(row=row, column=4, value=detail).font = Font(name='Arial', size=9)
        for col in range(1, 5):
            ws3.cell(row=row, column=col).border = border
        row += 1

for sheet in [ws, ws2, ws3]:
    sheet.freeze_panes = 'A2'

output = r'C:\Users\EliteBook\Downloads\Totales_Febrero_2026.xlsx'
wb.save(output)
print(f'Guardado en: {output}')
