"""
exporter.py — Generates the output Excel with 3 sheets.

Sheet 1: "Ventas por Período"
Sheet 2: "Consumos y Ventas s-Peso"
Sheet 3: "Observaciones"
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from models import PeriodoResult, FlavorPeriodoResult, AnomalyInfo

# ── Color palette ─────────────────────────────────────────────────────────────
C_HEADER_A      = "1F4E79"   # dark blue  (turno A headers)
C_HEADER_B      = "1E6B3C"   # dark green (turno B headers)
C_HEADER_CALC   = "404040"   # dark gray  (calculation headers)
C_WHITE         = "FFFFFF"
C_LIGHT_BLUE    = "DCE9F5"   # alternate row color
C_ANOMALY_ROW   = "FFE0E0"   # red - anomaly flavor row
C_COMMENT_ROW   = "FFF0F0"   # pink - comment/anomaly detail row
C_NEGATIVE      = "FF9900"   # orange bg for negative venta
C_POSITIVE_BLUE = "DCE9F5"   # light blue for positive venta
C_ENTRANTE_BG   = "E2EFDA"   # green for entrante rows
C_ENTRANTE_FG   = "375623"
C_LATA_BG       = "FFFF99"   # yellow for latas abiertas / ajuste
C_LATA_FG       = "7F6000"
C_EXTENDED      = "FFE5CC"   # orange for extended period header
C_SUBTOTAL_BG   = "E2EFDA"
C_CONSUMO_BG    = "FFE0E0"
C_OBS_STOCK     = "DCE9F5"
C_OBS_TEXT      = "FFFACD"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_WHITE, size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_border() -> Border:
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


# ── Sheet 1: Ventas por Período ───────────────────────────────────────────────

def _write_ventas(ws, periods: list):
    row = 1

    for periodo in periods:
        max_cerr_a = max((len(r.a_cerradas) for r in periodo.flavors.values()), default=0)
        max_cerr_b = max((len(r.b_cerradas) for r in periodo.flavors.values()), default=0)
        max_ent_a  = max((len(r.a_entrantes) for r in periodo.flavors.values()), default=0)

        # Column layout:
        # A: Sabor
        # B: Abierta_A | C: Celiaca_A | D..D+n: Cerradas_A | E: Entrantes_A | F: Total_A
        # G: Abierta_B | H: Celiaca_B | I..I+n: Cerradas_B | J: Entrantes_B | K: Total_B
        # L: Latas Abiertas | M: Ajuste | N: Dif. Bruta | O: Venta Neta (g)

        col_a_start   = 1    # Sabor
        col_a_abierta = 2
        col_a_celiaca = 3
        col_a_cerr1   = 4
        col_a_cerr_n  = 4 + max_cerr_a - 1 if max_cerr_a else 4
        col_a_ent     = col_a_cerr_n + 1 if max_cerr_a else 4
        col_a_total   = col_a_ent + 1

        col_b_abierta = col_a_total + 1
        col_b_celiaca = col_b_abierta + 1
        col_b_cerr1   = col_b_celiaca + 1
        col_b_cerr_n  = col_b_cerr1 + max_cerr_b - 1 if max_cerr_b else col_b_cerr1
        col_b_ent     = col_b_cerr_n + 1 if max_cerr_b else col_b_cerr1
        col_b_total   = col_b_ent + 1

        col_latas_ab  = col_b_total + 1
        col_ajuste    = col_latas_ab + 1
        col_dif_bruta = col_ajuste + 1
        col_venta_net = col_dif_bruta + 1

        total_cols = col_venta_net

        # ── Period header ─────────────────────────────────────────────────────
        bg = C_EXTENDED if periodo.extended else "2E75B6"
        header_text = f"{periodo.turno_a_name}  →  {periodo.turno_b_name}"
        if periodo.extended:
            header_text += "  ⚠ PERÍODO EXTENDIDO"

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        cell = ws.cell(row, 1, header_text)
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, size=12)
        cell.alignment = _align()
        row += 1

        # ── Sub-header ────────────────────────────────────────────────────────
        def hdr(c, txt, bg_hex, fg_hex=C_WHITE):
            cel = ws.cell(row, c, txt)
            cel.fill = _fill(bg_hex)
            cel.font = _font(bold=True, color=fg_hex, size=9)
            cel.alignment = _align()
            cel.border = _thin_border()

        hdr(1, "SABOR", C_HEADER_A)
        hdr(col_a_abierta, "ABIERTA", C_HEADER_A)
        hdr(col_a_celiaca, "CELIACA", C_HEADER_A)
        for i in range(max_cerr_a):
            hdr(col_a_cerr1 + i, f"CERR.{i+1}", C_HEADER_A)
        hdr(col_a_ent, "ENTRANT.", C_HEADER_A)
        hdr(col_a_total, "TOTAL A", C_HEADER_A)
        hdr(col_b_abierta, "ABIERTA", C_HEADER_B)
        hdr(col_b_celiaca, "CELIACA", C_HEADER_B)
        for i in range(max_cerr_b):
            hdr(col_b_cerr1 + i, f"CERR.{i+1}", C_HEADER_B)
        hdr(col_b_ent, "ENTRANT.", C_HEADER_B)
        hdr(col_b_total, "TOTAL B", C_HEADER_B)
        hdr(col_latas_ab, "LATAS AB.", C_HEADER_CALC)
        hdr(col_ajuste, "AJUSTE", C_HEADER_CALC)
        hdr(col_dif_bruta, "DIF.BRUTA", C_HEADER_CALC)
        hdr(col_venta_net, "VENTA NETA (g)", C_HEADER_CALC)
        row += 1

        # ── Data rows ─────────────────────────────────────────────────────────
        sorted_flavors = sorted(periodo.flavors.values(), key=lambda r: r.name)
        alternate = False

        for res in sorted_flavors:
            has_anomaly = bool(res.anomalies)
            row_bg = C_ANOMALY_ROW if has_anomaly else (C_LIGHT_BLUE if alternate else C_WHITE)
            alternate = not alternate

            def dval(c, val, fmt=None, bg=row_bg, fg="000000", bold=False):
                cel = ws.cell(row, c)
                if val is not None:
                    cel.value = val
                cel.fill = _fill(bg)
                cel.font = _font(bold=bold, color=fg, size=9)
                cel.alignment = _align("right") if isinstance(val, (int, float)) else _align("left")
                cel.border = _thin_border()

            dval(1, res.name, bold=has_anomaly)
            dval(col_a_abierta, res.a_abierta or None)
            dval(col_a_celiaca, res.a_celiaca or None)
            for i, v in enumerate(res.a_cerradas):
                dval(col_a_cerr1 + i, v)
            ent_a_sum = sum(res.a_entrantes) or None
            dval(col_a_ent, ent_a_sum, bg=C_ENTRANTE_BG if ent_a_sum else row_bg,
                 fg=C_ENTRANTE_FG if ent_a_sum else "000000")
            dval(col_a_total, res.a_total or None)

            dval(col_b_abierta, res.b_abierta or None)
            dval(col_b_celiaca, res.b_celiaca or None)
            for i, v in enumerate(res.b_cerradas):
                dval(col_b_cerr1 + i, v)
            ent_b_sum = sum(res.b_entrantes) or None
            dval(col_b_ent, ent_b_sum, bg=C_ENTRANTE_BG if ent_b_sum else row_bg,
                 fg=C_ENTRANTE_FG if ent_b_sum else "000000")
            dval(col_b_total, res.b_total or None)

            dval(col_latas_ab, res.latas_abiertas or None,
                 bg=C_LATA_BG if res.latas_abiertas else row_bg,
                 fg=C_LATA_FG if res.latas_abiertas else "000000")
            dval(col_ajuste, -res.ajuste if res.ajuste else None,
                 bg=C_LATA_BG if res.ajuste else row_bg,
                 fg=C_LATA_FG if res.ajuste else "000000")
            dval(col_dif_bruta, round(res.a_total - res.b_total, 0) or None)

            # Venta neta formatting
            vn = round(res.venta_neta, 0)
            if vn < 0:
                vn_bg, vn_fg = C_NEGATIVE, "CC0000"
            elif vn == 0:
                vn_bg, vn_fg = row_bg, "000000"
            else:
                vn_bg, vn_fg = C_POSITIVE_BLUE, "1F4E79"
            dval(col_venta_net, vn, bg=vn_bg, fg=vn_fg, bold=(vn < 0 or vn > 0))

            row += 1

            # Anomaly comment rows
            for anom in res.anomalies:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
                cel = ws.cell(row, 1, anom.message)
                cel.fill = _fill(C_COMMENT_ROW)
                cel.font = Font(italic=True, color="CC0000", size=8, name="Arial")
                cel.alignment = _align("left", wrap=True)
                row += 1

        # ── Subtotal row ──────────────────────────────────────────────────────
        total_venta = sum(
            round(r.venta_neta, 0)
            for r in periodo.flavors.values()
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_venta_net - 1)
        c = ws.cell(row, 1, "SUBTOTAL DEL PERÍODO")
        c.fill = _fill(C_SUBTOTAL_BG)
        c.font = _font(bold=True, color="1E6B3C", size=10)
        c.alignment = _align("right")

        c2 = ws.cell(row, col_venta_net, total_venta)
        c2.fill = _fill(C_SUBTOTAL_BG)
        c2.font = _font(bold=True, color="1E6B3C", size=10)
        c2.alignment = _align("right")
        row += 2  # blank line between periods

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 28
    for col in range(2, 30):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.freeze_panes = "B3"


# ── Sheet 2: Consumos y Ventas s-Peso ────────────────────────────────────────

def _write_consumos(ws, periods: list):
    row = 1

    # Section: Ventas sin peso
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    h = ws.cell(row, 1, "VENTAS DESPUÉS DEL PESO (por período)")
    h.fill = _fill("1E6B3C")
    h.font = _font(bold=True, size=11)
    h.alignment = _align()
    row += 1

    for col, title in enumerate(["PERÍODO", "TURNO", "DESCRIPCIÓN", "GRAMOS"], 1):
        c = ws.cell(row, col, title)
        c.fill = _fill(C_HEADER_B)
        c.font = _font(bold=True, size=9)
        c.alignment = _align()
    row += 1

    for periodo in periods:
        for entry in periodo.ventas_sin_peso:
            ws.cell(row, 1, f"{periodo.turno_a_name} → {periodo.turno_b_name}")
            ws.cell(row, 2, "")
            ws.cell(row, 3, entry.raw_text)
            ws.cell(row, 4, entry.grams)
            for c in range(1, 5):
                ws.cell(row, c).fill = _fill(C_ENTRANTE_BG)
                ws.cell(row, c).font = Font(size=9, name="Arial")
            row += 1

    row += 1

    # Section: Consumos internos
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    h2 = ws.cell(row, 1, "CONSUMO INTERNO (empleados)")
    h2.fill = _fill("C00000")
    h2.font = _font(bold=True, size=11)
    h2.alignment = _align()
    row += 1

    for col, title in enumerate(["PERÍODO", "EMPLEADO", "DESCRIPCIÓN", "GRAMOS"], 1):
        c = ws.cell(row, col, title)
        c.fill = _fill("C00000")
        c.font = _font(bold=True, size=9)
        c.alignment = _align()
    row += 1

    total_consumo = 0.0
    for periodo in periods:
        for entry in periodo.consumos:
            ws.cell(row, 1, f"{periodo.turno_a_name} → {periodo.turno_b_name}")
            ws.cell(row, 2, entry.employee)
            ws.cell(row, 3, entry.description)
            ws.cell(row, 4, entry.grams)
            total_consumo += entry.grams
            for c in range(1, 5):
                ws.cell(row, c).fill = _fill(C_CONSUMO_BG)
                ws.cell(row, c).font = Font(size=9, name="Arial")
            row += 1

    # Total
    row += 1
    ws.cell(row, 3, "TOTAL CONSUMO INTERNO (g)").font = _font(bold=True, color="000000", size=10)
    t = ws.cell(row, 4, total_consumo)
    t.font = _font(bold=True, color="000000", size=10)
    t.fill = _fill(C_LATA_BG)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12


# ── Sheet 3: Observaciones ────────────────────────────────────────────────────

def _write_observaciones(ws, periods: list):
    row = 1

    headers = ["DÍA / PERÍODO", "SABOR / ÍTEM", "ABIERTO (g)", "CERRADO (g)", "ENTRANTES / NOTA"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.fill = _fill(C_HEADER_A)
        c.font = _font(bold=True, size=10)
        c.alignment = _align()
    row += 1

    for periodo in periods:
        label = f"{periodo.turno_b_name}"
        for obs in periodo.observaciones:
            is_stock = obs.abierta is not None or obs.cerrada is not None or obs.entrante is not None
            bg = C_OBS_STOCK if is_stock else C_OBS_TEXT

            ws.cell(row, 1, label)
            ws.cell(row, 2, obs.flavor)
            ws.cell(row, 3, obs.abierta)
            ws.cell(row, 4, obs.cerrada)
            ws.cell(row, 5, obs.entrante or obs.nota)

            for c in range(1, 6):
                cel = ws.cell(row, c)
                cel.fill = _fill(bg)
                if not is_stock:
                    cel.font = Font(italic=True, size=9, name="Arial")
                else:
                    cel.font = Font(size=9, name="Arial")
                cel.alignment = Alignment(horizontal="center", vertical="center")
            row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 30


# ── Main export function ──────────────────────────────────────────────────────

def export(periods: list, output_path: str):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Ventas por Período"
    _write_ventas(ws1, periods)

    ws2 = wb.create_sheet("Consumos y Ventas s-Peso")
    _write_consumos(ws2, periods)

    ws3 = wb.create_sheet("Observaciones")
    _write_observaciones(ws3, periods)

    wb.save(output_path)
    print(f"OK Guardado: {output_path}")
