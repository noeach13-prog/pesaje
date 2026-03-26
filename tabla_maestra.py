"""
tabla_maestra.py - Genera tabla maestra normalizada del workbook.
Lee el Excel crudo y produce una fila por sabor por turno.
NO infiere, NO corrige, solo normaliza y reporta observaciones.
"""
import sys
import re
import json
import datetime
import openpyxl

INPUT = r'C:\Users\EliteBook\Downloads\Febrero San Martin 2026 (1).xlsx'
OUTPUT_XLSX = r'C:\Users\EliteBook\Pesaje\tabla_maestra.xlsx'
OUTPUT_JSON = r'C:\Users\EliteBook\Pesaje\tabla_maestra.json'

# ── Normalization helpers ──

# Known aliases: raw_upper -> canonical
_NAME_ALIASES = {
    'TIRAMIZU': 'TIRAMISU',
}

# Accent map
_ACCENT_MAP = str.maketrans('AEIOUUN', 'AEIOUUN')
_ACCENT_MAP_FULL = str.maketrans(
    '\u00c1\u00c9\u00cd\u00d3\u00da\u00dc\u00d1',  # accented
    'AEIOUUN'
)

def normalize_name(name):
    """Normalize flavor name: uppercase, remove accents, collapse spaces."""
    if name is None:
        return ''
    s = str(name).strip().upper()
    s = s.translate(_ACCENT_MAP_FULL)
    s = re.sub(r'\s+', ' ', s)
    return s

def safe_float(val):
    """Convert cell value to float, return None if not numeric."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, datetime.datetime):
        return None  # Excel date misparse
    if isinstance(val, str):
        s = val.strip().replace(',', '.')
        if s in ('-', ''):
            return None
        if s == '0':
            return 0.0
        try:
            return float(s)
        except ValueError:
            return None
    return None

def classify_cell(val, col_letter, expected_type):
    """Classify a cell: returns (numeric_value_or_None, observation_string).
    observation_string describes anomalies found in the cell."""
    obs = []

    if val is None:
        return None, None  # truly empty

    if isinstance(val, bool):
        obs.append(f"col_{col_letter}: boolean={val}")
        return None, '; '.join(obs)

    if isinstance(val, datetime.datetime):
        obs.append(f"col_{col_letter}: datetime={val} (probable misparse)")
        return None, '; '.join(obs)

    if isinstance(val, str):
        s = val.strip()
        if s == '':
            return None, None  # empty string = empty
        if s == '-':
            return None, f"col_{col_letter}: guion '-' (vacio explicito)"
        if s == '0':
            return 0.0, None
        # Try numeric
        try:
            v = float(s.replace(',', '.'))
            obs.append(f"col_{col_letter}: texto_numerico='{s}'")
            return v, '; '.join(obs)
        except ValueError:
            obs.append(f"col_{col_letter}: texto_no_numerico='{s}'")
            return None, '; '.join(obs)

    if isinstance(val, (int, float)):
        v = float(val)
        if v < 0:
            obs.append(f"col_{col_letter}: negativo={v}")
        if v == 0:
            return 0.0, None
        return v, '; '.join(obs) if obs else None

    obs.append(f"col_{col_letter}: tipo_inesperado={type(val).__name__}={val}")
    return None, '; '.join(obs)


# ── Parse the workbook ──

def extract_fecha_turno(sheet_name):
    """Extract fecha and turno from sheet name like 'Sabado 1 (DIA)'."""
    name = sheet_name.strip()

    # Extract turno
    turno = None
    m_turno = re.search(r'\((DIA|NOCHE)\)', name, re.IGNORECASE)
    if m_turno:
        turno = m_turno.group(1).upper()

    # Extract day number
    m_day = re.search(r'(\d+)', name)
    dia = int(m_day.group(1)) if m_day else None

    # Extract day of week
    m_dow = re.match(r'([A-Za-z\u00e1\u00e9\u00ed\u00f3\u00fa]+)', name)
    dia_semana = m_dow.group(1) if m_dow else None

    # Build fecha
    if dia:
        fecha = f"2026-02-{dia:02d}"
    else:
        fecha = None

    return fecha, turno, dia_semana


def build_master_table(path):
    """Read the full workbook and return a list of row dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    all_names_raw = {}  # normalized -> set of originals

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]

        # Skip non-data sheets
        a1 = ws['A1'].value
        if a1 != 'SABORES':
            continue

        fecha, turno, dia_semana = extract_fecha_turno(sheet_name)
        max_row = ws.max_row

        # Find POSTRES marker and flavor rows
        for r in range(2, max_row + 1):
            a_val = ws.cell(r, 1).value

            # Stop at POSTRES
            if a_val is not None and isinstance(a_val, str):
                if normalize_name(a_val).startswith('POSTRES'):
                    break

            # Skip if col A is None (totals row or gap)
            if a_val is None:
                continue

            # Skip header repeat
            if normalize_name(a_val) == 'SABORES':
                continue

            # ── Extract flavor data ──
            sabor_original = str(a_val).strip() if a_val else ''
            sabor_norm = normalize_name(a_val)

            if not sabor_norm:
                continue

            # Track all name variants
            if sabor_norm not in all_names_raw:
                all_names_raw[sabor_norm] = set()
            all_names_raw[sabor_norm].add(sabor_original)

            # ── Read each cell individually with classification ──
            observations = []

            # Col B = Abierta
            ab_raw = ws.cell(r, 2).value
            abierta, ab_obs = classify_cell(ab_raw, 'B', 'abierta')
            if ab_obs:
                observations.append(ab_obs)

            # Col C = Celiaca
            cel_raw = ws.cell(r, 3).value
            celiaca, cel_obs = classify_cell(cel_raw, 'C', 'celiaca')
            if cel_obs:
                observations.append(cel_obs)

            # Cols D-I = Cerradas
            cerradas = []
            cerradas_detail = []  # for observations
            for c in range(4, 10):
                col_letter = chr(64 + c)
                raw_val = ws.cell(r, c).value
                val, obs = classify_cell(raw_val, col_letter, 'cerrada')
                if obs:
                    observations.append(obs)
                if val is not None and val > 0:
                    cerradas.append(val)
                    cerradas_detail.append({'col': col_letter, 'val': val})
                elif val is not None and val == 0:
                    # Zero in cerrada slot - note it
                    observations.append(f"col_{col_letter}: cerrada=0 (slot con cero)")
                elif val is not None and val < 0:
                    cerradas.append(val)  # keep negative, flag it
                    cerradas_detail.append({'col': col_letter, 'val': val})
                    observations.append(f"col_{col_letter}: cerrada_negativa={val}")

            # Cols J-K = Entrantes
            entrantes = []
            entrantes_detail = []
            for c in range(10, 12):
                col_letter = chr(64 + c)
                raw_val = ws.cell(r, c).value
                val, obs = classify_cell(raw_val, col_letter, 'entrante')
                if obs:
                    observations.append(obs)
                if val is not None and val > 0:
                    entrantes.append(val)
                    entrantes_detail.append({'col': col_letter, 'val': val})
                elif val is not None and val == 0:
                    observations.append(f"col_{col_letter}: entrante=0 (slot con cero)")
                elif val is not None and val < 0:
                    entrantes.append(val)
                    entrantes_detail.append({'col': col_letter, 'val': val})
                    observations.append(f"col_{col_letter}: entrante_negativa={val}")

            # Determine if row is fully empty
            has_any_data = (
                (abierta is not None and abierta != 0) or
                (celiaca is not None and celiaca != 0) or
                len(cerradas) > 0 or
                len(entrantes) > 0
            )

            if not has_any_data:
                observations.append("fila_sin_datos_numericos")

            # Check alias
            alias_propuesto = _NAME_ALIASES.get(sabor_norm, None)
            if alias_propuesto and alias_propuesto != sabor_norm:
                observations.append(f"alias_propuesto: {sabor_norm} -> {alias_propuesto}")

            row = {
                'fecha': fecha,
                'turno': turno,
                'sabor_original': sabor_original,
                'sabor_normalizado': sabor_norm,
                'abierta': abierta if abierta is not None else 0.0,
                'celiaca': celiaca if celiaca is not None else 0.0,
                'cerradas': cerradas,
                'entrantes': entrantes,
                'hoja_origen': sheet_name.strip(),
                'fila_origen': r,
                'observaciones_de_lectura': '; '.join(observations) if observations else '',
            }
            rows.append(row)

    wb.close()
    return rows, all_names_raw


def detect_alias_candidates(all_names_raw):
    """Find names that look like variants of each other."""
    from difflib import SequenceMatcher

    norms = sorted(all_names_raw.keys())
    candidates = []

    for i, n1 in enumerate(norms):
        for n2 in norms[i+1:]:
            # Skip if same
            if n1 == n2:
                continue

            # Levenshtein-like ratio
            ratio = SequenceMatcher(None, n1, n2).ratio()
            if ratio >= 0.75 and ratio < 1.0:
                candidates.append({
                    'nombre_1': n1,
                    'nombre_2': n2,
                    'similitud': round(ratio, 3),
                    'variantes_1': sorted(all_names_raw[n1]),
                    'variantes_2': sorted(all_names_raw[n2]),
                })

            # Also check if one is prefix/suffix of other
            if n1 in n2 or n2 in n1:
                if ratio < 0.75:  # not already caught
                    candidates.append({
                        'nombre_1': n1,
                        'nombre_2': n2,
                        'similitud': round(ratio, 3),
                        'variantes_1': sorted(all_names_raw[n1]),
                        'variantes_2': sorted(all_names_raw[n2]),
                        'nota': 'substring match',
                    })

    return candidates


def write_xlsx(rows, alias_candidates, output_path):
    """Write the master table to Excel."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Master table ──
    ws = wb.active
    ws.title = "Tabla Maestra"

    headers = [
        'fecha', 'turno', 'sabor_original', 'sabor_normalizado',
        'abierta', 'celiaca', 'cerradas', 'entrantes',
        'hoja_origen', 'fila_origen', 'observaciones_de_lectura'
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)

    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, row['fecha'])
        ws.cell(i, 2, row['turno'])
        ws.cell(i, 3, row['sabor_original'])
        ws.cell(i, 4, row['sabor_normalizado'])
        ws.cell(i, 5, row['abierta'])
        ws.cell(i, 6, row['celiaca'])
        ws.cell(i, 7, json.dumps(row['cerradas']))
        ws.cell(i, 8, json.dumps(row['entrantes']))
        ws.cell(i, 9, row['hoja_origen'])
        ws.cell(i, 10, row['fila_origen'])
        ws.cell(i, 11, row['observaciones_de_lectura'])

    # Auto-width (approximate)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64+c) if c < 27 else 'A'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['K'].width = 60

    # ── Sheet 2: Alias candidates ──
    ws2 = wb.create_sheet("Alias Candidatos")
    ws2.cell(1, 1, "nombre_1")
    ws2.cell(1, 2, "nombre_2")
    ws2.cell(1, 3, "similitud")
    ws2.cell(1, 4, "variantes_1")
    ws2.cell(1, 5, "variantes_2")
    ws2.cell(1, 6, "nota")

    for i, ac in enumerate(alias_candidates, 2):
        ws2.cell(i, 1, ac['nombre_1'])
        ws2.cell(i, 2, ac['nombre_2'])
        ws2.cell(i, 3, ac['similitud'])
        ws2.cell(i, 4, ', '.join(ac['variantes_1']))
        ws2.cell(i, 5, ', '.join(ac['variantes_2']))
        ws2.cell(i, 6, ac.get('nota', ''))

    # ── Sheet 3: Summary stats ──
    ws3 = wb.create_sheet("Resumen")
    ws3.cell(1, 1, "Total filas")
    ws3.cell(1, 2, len(rows))
    ws3.cell(2, 1, "Hojas procesadas")
    hojas = sorted(set(r['hoja_origen'] for r in rows))
    ws3.cell(2, 2, len(hojas))
    ws3.cell(3, 1, "Sabores unicos (normalizado)")
    sabores_unicos = sorted(set(r['sabor_normalizado'] for r in rows))
    ws3.cell(3, 2, len(sabores_unicos))
    ws3.cell(4, 1, "Filas con observaciones")
    ws3.cell(4, 2, sum(1 for r in rows if r['observaciones_de_lectura']))
    ws3.cell(5, 1, "Filas sin datos numericos")
    ws3.cell(5, 2, sum(1 for r in rows if 'fila_sin_datos_numericos' in r['observaciones_de_lectura']))
    ws3.cell(6, 1, "Alias candidatos")
    ws3.cell(6, 2, len(alias_candidates))

    # List all sheets
    ws3.cell(8, 1, "Hojas procesadas:")
    for i, h in enumerate(hojas):
        n_rows = sum(1 for r in rows if r['hoja_origen'] == h)
        ws3.cell(9 + i, 1, h)
        ws3.cell(9 + i, 2, n_rows)

    wb.save(output_path)


def main():
    print("Leyendo workbook...")
    rows, all_names_raw = build_master_table(INPUT)

    print(f"Total filas extraidas: {len(rows)}")
    print(f"Sabores unicos (normalizado): {len(set(r['sabor_normalizado'] for r in rows))}")
    print(f"Hojas procesadas: {len(set(r['hoja_origen'] for r in rows))}")

    # ── Alias detection ──
    alias_candidates = detect_alias_candidates(all_names_raw)

    # ── Muestra de 20 filas ──
    print("\n" + "=" * 140)
    print("MUESTRA DE 20 FILAS:")
    print("=" * 140)

    # Take first 10 and last 10 for variety
    sample = rows[:10] + rows[-10:]

    header = f"{'fecha':<12} {'turno':<6} {'sabor_orig':<22} {'sabor_norm':<22} {'ab':>7} {'cel':>5} {'cerradas':<25} {'entrantes':<18} {'hoja':<28} {'fila':>4} {'obs'}"
    print(header)
    print("-" * 200)

    for r in sample:
        cerr_str = ','.join(f'{c:.0f}' for c in r['cerradas'])
        ent_str = ','.join(f'{e:.0f}' for e in r['entrantes'])
        obs_short = r['observaciones_de_lectura'][:50] if r['observaciones_de_lectura'] else ''
        print(f"{r['fecha'] or '???':<12} {r['turno'] or '???':<6} {r['sabor_original']:<22} {r['sabor_normalizado']:<22} {r['abierta']:>7.0f} {r['celiaca']:>5.0f} [{cerr_str:<23}] [{ent_str:<16}] {r['hoja_origen']:<28} {r['fila_origen']:>4} {obs_short}")

    # ── Stats ──
    print(f"\n{'='*140}")
    print("ESTADISTICAS:")
    print(f"{'='*140}")

    filas_con_obs = [r for r in rows if r['observaciones_de_lectura']]
    filas_sin_datos = [r for r in rows if 'fila_sin_datos_numericos' in r['observaciones_de_lectura']]

    print(f"  Total filas:                {len(rows)}")
    print(f"  Filas con observaciones:    {len(filas_con_obs)}")
    print(f"  Filas sin datos numericos:  {len(filas_sin_datos)}")
    print(f"  Alias candidatos:           {len(alias_candidates)}")

    # ── Alias candidates ──
    if alias_candidates:
        print(f"\n{'='*140}")
        print("ALIAS CANDIDATOS (similitud >= 0.75 o substring):")
        print(f"{'='*140}")
        for ac in alias_candidates:
            nota = f" [{ac['nota']}]" if ac.get('nota') else ''
            print(f"  {ac['nombre_1']:<25} <-> {ac['nombre_2']:<25} sim={ac['similitud']:.3f}{nota}")
            print(f"    variantes_1: {', '.join(ac['variantes_1'])}")
            print(f"    variantes_2: {', '.join(ac['variantes_2'])}")

    # ── Name variants ──
    multi_variant = {k: v for k, v in all_names_raw.items() if len(v) > 1}
    if multi_variant:
        print(f"\n{'='*140}")
        print("SABORES CON MULTIPLES VARIANTES DE ESCRITURA:")
        print(f"{'='*140}")
        for norm, variants in sorted(multi_variant.items()):
            print(f"  {norm}: {sorted(variants)}")

    # ── All observations ──
    if filas_con_obs:
        print(f"\n{'='*140}")
        print(f"TODAS LAS OBSERVACIONES DE LECTURA ({len(filas_con_obs)} filas):")
        print(f"{'='*140}")
        for r in filas_con_obs:
            print(f"  {r['hoja_origen']:<28} fila={r['fila_origen']:>3} {r['sabor_normalizado']:<20} | {r['observaciones_de_lectura']}")

    # ── Sabor count per shift ──
    print(f"\n{'='*140}")
    print("SABORES POR HOJA:")
    print(f"{'='*140}")
    from collections import Counter
    hojas = sorted(set(r['hoja_origen'] for r in rows), key=lambda h: next(r['fila_origen'] for r in rows if r['hoja_origen'] == h))
    # Actually sort by first occurrence in the workbook
    hoja_order = {}
    for r in rows:
        if r['hoja_origen'] not in hoja_order:
            hoja_order[r['hoja_origen']] = len(hoja_order)
    hojas = sorted(set(r['hoja_origen'] for r in rows), key=lambda h: hoja_order[h])

    for h in hojas:
        h_rows = [r for r in rows if r['hoja_origen'] == h]
        n_with_data = sum(1 for r in h_rows if 'fila_sin_datos_numericos' not in r['observaciones_de_lectura'])
        n_without = sum(1 for r in h_rows if 'fila_sin_datos_numericos' in r['observaciones_de_lectura'])
        obs_flag = f" ({n_without} sin datos)" if n_without > 0 else ""
        print(f"  {h:<35} {n_with_data:>3} sabores con datos{obs_flag}")

    # ── Write outputs ──
    print(f"\nEscribiendo Excel -> {OUTPUT_XLSX}")
    write_xlsx(rows, alias_candidates, OUTPUT_XLSX)

    print(f"Escribiendo JSON -> {OUTPUT_JSON}")
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump({
            'rows': rows,
            'alias_candidates': alias_candidates,
            'name_variants': {k: sorted(v) for k, v in all_names_raw.items()},
        }, f, ensure_ascii=False, indent=2)

    print("\nDone.")


if __name__ == '__main__':
    main()
