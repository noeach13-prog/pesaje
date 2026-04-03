"""
exporter_multi.py -- Excel de reporte unificado, diseñado para lectura humana.

Hojas:
  - "Guía":        leyenda de colores y glosario
  - "Resumen":     tabla mensual con hipervínculos a cada día
  - "Día 5", ...:  detalle por día (5 columnas, lenguaje simple, VDP incluido)
  - "Detalle técnico": trazabilidad completa para análisis avanzado
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1 as FORMAT_NUMBER_COMMA_SEP1
from .modelos import (
    ResultadoC3, SaborClasificado, StatusC3,
    Correccion, Banda, ResultadoDia, TipoResolucion,
)
from typing import List, Tuple, Optional


# ===================================================================
# PALETA
# ===================================================================
C_HEADER_DARK   = "1F4E79"   # azul oscuro — encabezados principales
C_HEADER_MID    = "2E75B6"   # azul medio — subencabezados
C_WHITE         = "FFFFFF"
C_ALT           = "F5F9FF"   # azul muy claro — filas alternas
C_LIMPIO        = "FFFFFF"   # sin cambios — blanco puro
C_AJUSTE_LEVE   = "E8F5E9"   # verde muy claro — ajuste pequeño
C_CORREGIDO     = "DDEEFF"   # azul claro — corregido con confianza
C_ESTIMADO      = "FFF8E1"   # amarillo claro — estimado / revisar
C_SOLO          = "F0F0F0"   # gris claro — solo un turno
C_TOTAL_ROW     = "E8F0FE"   # azul pálido — fila de totales del día
C_VDP_ROW       = "F3E5F5"   # lila muy claro — fila VDP
C_TAPAS_ROW     = "FFF3E0"   # naranja muy claro — fila tapas
C_GRAND_TOTAL   = "1F4E79"   # azul oscuro — fila gran total
C_ACCENT        = "FF6B35"   # naranja acento — para destacar


# ===================================================================
# HELPERS DE ESTILO
# ===================================================================

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)

def _align(h="center", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _border(style='thin', color='D0D0D0'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color='AAAAAA'):
    s = Side(style='medium', color=color)
    n = Side(style=None)
    return Border(bottom=s, left=n, right=n, top=n)

def _cell(ws, row, col, value=None, font=None, fill=None,
          align=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    if number_format: c.number_format = number_format
    return c


# ===================================================================
# SITUACIÓN — traducción de estados técnicos a lenguaje humano
# ===================================================================

_SITUACION_PRIORITY = {
    '! Corregido':      0,
    '~ Estimado':       1,
    '~ Ajuste leve':    2,
    'OK Sin cambios':   3,
    '? Sin resolver':   3,
    '- Solo un turno':  4,
}

def _clasificar_fila(sc: SaborClasificado, corr: Optional[Correccion], venta_raw_fallback: int = 0):
    """Devuelve (situacion_str, color_hex, delta_display, venta_final)."""
    # SOLO un turno
    if sc.status in (StatusC3.SOLO_DIA, StatusC3.SOLO_NOCHE):
        vf = 0
        return '- Solo un turno', C_SOLO, None, vf

    # Prototipo C3 (PF) sin corrección C4
    proto = sc.prototipo
    if corr is None and proto is None:
        if sc.venta_final_c3 is not None:
            return 'OK Sin cambios', C_LIMPIO, None, sc.venta_final_c3
        # ESCALAR_C4: venta_final_c3 no resuelta, mostrar raw para coherencia de columna
        return '? Sin resolver', C_ESTIMADO, None, venta_raw_fallback

    # Usar corrección C4 si existe, si no el prototipo
    if corr is not None:
        delta = corr.delta
        vf    = corr.venta_corregida
        banda = corr.banda.value
        tipo_res = corr.tipo_resolucion.value
    else:
        delta = proto.delta
        vf    = proto.venta_corregida
        banda = 'CONFIRMADO' if proto.confianza >= 0.85 else 'ESTIMADO'
        tipo_res = 'RESUELTO_INDIVIDUAL'

    if tipo_res == 'FORZADO_H0':
        return '~ Estimado', C_ESTIMADO, delta, vf

    if abs(delta) <= 200:
        return '~ Ajuste leve', C_AJUSTE_LEVE, delta, vf

    if banda == 'CONFIRMADO':
        return '! Corregido', C_CORREGIDO, delta, vf

    return '~ Estimado', C_ESTIMADO, delta, vf


def _nota_legible(sc: SaborClasificado, corr: Optional[Correccion]) -> str:
    """Convierte motivo técnico en frase corta en español."""
    proto = sc.prototipo

    if corr is not None:
        m = corr.motivo
        if 'COMPOSICION' in m:
            try:
                nx = m.split('COMPOSICION')[1].split('x')[0].strip()
                return f'Corrección múltiple ({nx} causas combinadas)'
            except Exception:
                return 'Corrección múltiple combinada'
        if 'FORZADO_H0' in m:
            return 'Estimación forzada — revisar manualmente'
        if 'PHANTOM_DIA' in m:
            return 'Lata fantasma eliminada (turno DÍA)'
        if 'PHANTOM_NOCHE' in m:
            return 'Lata fantasma eliminada (turno NOCHE)'
        if 'OMISION_DIA' in m:
            return 'Lata omitida en DÍA recuperada'
        if 'OMISION_NOCHE' in m:
            return 'Lata omitida en NOCHE recuperada'
        if 'MISMATCH_LEVE' in m:
            return 'Varianza de pesaje ajustada'
        if 'GENEALOGIA_ENT_CERR' in m:
            return 'Entrante reclasificado como lata'
        if 'ENTRANTE_MISMO_CAN' in m:
            return 'Entrante duplicado detectado'
        if 'DUPLICADO_CERRADA' in m:
            return 'Lata duplicada detectada'
        return ''

    if proto is not None:
        notas_pf = {
            'PF1': 'Error de dígito corregido',
            'PF2': 'Entrante duplicado DÍA→NOCHE',
            'PF3': 'Phantom RM-3 detectado',
            'PF4': 'Lata cerrada omitida en NOCHE',
            'PF5': 'Lata cerrada omitida en DÍA',
            'PF6': 'Apertura + phantom combinados',
            'PF7': 'Abierta imposible corregida',
        }
        return notas_pf.get(proto.codigo, proto.codigo)

    return ''


# ===================================================================
# HOJA 1: GUÍA DE LECTURA
# ===================================================================

def _write_guia(ws):
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 58
    ws.sheet_view.showGridLines = False

    row = 1

    # Título
    ws.merge_cells(f'B{row}:C{row}')
    _cell(ws, row, 2, 'REPORTE DE VENTAS — GUÍA DE LECTURA',
          font=_font(bold=True, color=C_WHITE, size=14),
          fill=_fill(C_HEADER_DARK),
          align=_align(h='left', indent=1))
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f'B{row}:C{row}')
    _cell(ws, row, 2,
          'Análisis automático de stock de helados — San Martín Febrero 2026',
          font=_font(italic=True, color='555555', size=10),
          fill=_fill('EEF4FF'),
          align=_align(h='left', indent=1))
    row += 2

    # Sección: qué es
    ws.merge_cells(f'B{row}:C{row}')
    _cell(ws, row, 2, 'QUÉ ES ESTE REPORTE',
          font=_font(bold=True, color=C_WHITE, size=10),
          fill=_fill(C_HEADER_MID),
          align=_align(h='left', indent=1))
    ws.row_dimensions[row].height = 20
    row += 1

    desc = (
        'Este archivo consolida las ventas de helado de cada día, comparando el '
        'stock del turno DÍA con el del turno NOCHE. Cuando el sistema detecta '
        'inconsistencias en los registros (latas fantasma, omisiones, errores de '
        'dígito), aplica correcciones automáticas y las documenta. La hoja '
        '"Detalle técnico" contiene la trazabilidad completa de cada corrección.'
    )
    ws.merge_cells(f'B{row}:C{row}')
    c = _cell(ws, row, 2, desc,
              font=_font(size=9),
              fill=_fill('F8FBFF'),
              align=_align(h='left', v='top', wrap=True, indent=1))
    ws.row_dimensions[row].height = 52
    row += 2

    # Sección: leyenda
    ws.merge_cells(f'B{row}:C{row}')
    _cell(ws, row, 2, 'LEYENDA DE COLORES',
          font=_font(bold=True, color=C_WHITE, size=10),
          fill=_fill(C_HEADER_MID),
          align=_align(h='left', indent=1))
    ws.row_dimensions[row].height = 20
    row += 1

    leyenda = [
        (C_LIMPIO,      'OK Sin cambios',    'El registro del día no presentó inconsistencias. Valor directo de la planilla.'),
        (C_AJUSTE_LEVE, '~ Ajuste leve',    'Varianza de pesaje menor (≤ 200 g). Corrección automática de pequeña medición.'),
        (C_CORREGIDO,   '! Corregido',      'Se detectó y corrigió un problema con alta confianza (lata fantasma, omisión, error de dígito, etc.).'),
        (C_ESTIMADO,    '~ Estimado',        'Corrección aplicada con confianza media o baja. Revisar si hay dudas sobre el registro.'),
        (C_SOLO,        '- Solo un turno',  'El sabor no tiene datos en ambos turnos del día. No se puede calcular venta.'),
    ]

    for color, label, texto in leyenda:
        _cell(ws, row, 2, label,
              font=_font(bold=True, size=9),
              fill=_fill(color),
              align=_align(h='left', indent=1),
              border=_border(color='CCCCCC'))
        _cell(ws, row, 3, texto,
              font=_font(size=9),
              fill=_fill(color),
              align=_align(h='left', wrap=True),
              border=_border(color='CCCCCC'))
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1

    # Sección: glosario
    ws.merge_cells(f'B{row}:C{row}')
    _cell(ws, row, 2, 'GLOSARIO',
          font=_font(bold=True, color=C_WHITE, size=10),
          fill=_fill(C_HEADER_MID),
          align=_align(h='left', indent=1))
    ws.row_dimensions[row].height = 20
    row += 1

    glosario = [
        ('Venta (g)',       'Gramos vendidos en el período DÍA → NOCHE. Es el valor que va a contabilidad.'),
        ('Ajuste (g)',      'Corrección detectada automáticamente. Vacío = sin cambios. +/- = gramos ajustados respecto al registro.'),
        ('VDP',             'Ventas de Postres — registradas por monto sin identificación de lata individual. Se suman al total del día.'),
        ('Aperturas',       'Descuento por tapas de latas abiertas durante el día (280 g por apertura).'),
        ('Total del día',   'Venta de helados (con ajustes) + VDP − descuento por aperturas.'),
    ]

    for termino, definicion in glosario:
        _cell(ws, row, 2, termino,
              font=_font(bold=True, size=9),
              fill=_fill('F8FBFF'),
              align=_align(h='left', indent=1),
              border=_border(color='DDDDDD'))
        _cell(ws, row, 3, definicion,
              font=_font(size=9),
              fill=_fill('F8FBFF'),
              align=_align(h='left', wrap=True),
              border=_border(color='DDDDDD'))
        ws.row_dimensions[row].height = 30
        row += 1


# ===================================================================
# HOJA 2: RESUMEN MENSUAL
# ===================================================================

def _write_resumen(ws, resultados, sheet_names: dict):
    """
    Una fila por día. El nombre del día es hipervínculo a la hoja correspondiente.
    sheet_names: {dia_label -> nombre_hoja_excel}
    """
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.sheet_view.showGridLines = False

    hdr_font   = _font(bold=True, color=C_WHITE, size=10)
    hdr_fill   = _fill(C_HEADER_DARK)
    hdr_align  = _align(wrap=True)
    brd        = _border()

    headers = ['Día', 'Venta helados (g)', 'VDP (g)', 'Aperturas (g)', 'Total del día (g)',
               'Correcciones', 'Sin datos']
    for col, h in enumerate(headers, 1):
        _cell(ws, 1, col, h, hdr_font, hdr_fill, hdr_align, brd)

    ws.row_dimensions[1].height = 28

    row = 2
    for datos, cont, c3, c4, resultado in resultados:
        label    = resultado.dia_label
        hoja     = sheet_names.get(label, f'Dia {label}')

        alt_fill = _fill(C_ALT if row % 2 == 0 else C_WHITE)

        # Celda del día con hipervínculo
        c = _cell(ws, row, 1, f'Día {label}',
                  font=Font(bold=True, color='0563C1', underline='single',
                            size=10, name='Calibri'),
                  fill=alt_fill, align=_align(), border=brd)
        c.hyperlink = f"#'{hoja}'!A1"

        n_sin_datos = sum(
            1 for sc in c3.sabores.values()
            if sc.status in (StatusC3.SOLO_DIA, StatusC3.SOLO_NOCHE)
        )

        vals = [
            resultado.venta_refinado,
            resultado.vdp,
            -resultado.lid_discount,
            resultado.total_refinado,
            len(c4.correcciones),
            n_sin_datos,
        ]
        for col, val in enumerate(vals, 2):
            f = _font(bold=(col == 5))   # negrita en Total
            fmt = '#,##0' if col <= 5 else None
            _cell(ws, row, col, val, f, alt_fill, _align(), brd,
                  number_format=fmt)
        row += 1

    # Fila totales
    row += 1
    _cell(ws, row, 1, 'TOTAL', _font(bold=True, color=C_WHITE),
          _fill(C_HEADER_DARK), _align(), brd)
    totales = [
        sum(r.venta_refinado    for _, _, _, _, r in resultados),
        sum(r.vdp               for _, _, _, _, r in resultados),
        -sum(r.lid_discount     for _, _, _, _, r in resultados),
        sum(r.total_refinado    for _, _, _, _, r in resultados),
        sum(len(c4.correcciones)for _, _, c3, c4, _ in resultados),
        sum(
            sum(1 for sc in c3.sabores.values()
                if sc.status in (StatusC3.SOLO_DIA, StatusC3.SOLO_NOCHE))
            for _, _, c3, _, _ in resultados
        ),
    ]
    for col, val in enumerate(totales, 2):
        _cell(ws, row, col, val,
              _font(bold=True, color=C_WHITE),
              _fill(C_HEADER_DARK), _align(), brd,
              number_format='#,##0')


# ===================================================================
# HOJA POR DÍA
# ===================================================================

def _write_dia(ws, label, c3, c4, resultado, cont=None):
    """Hoja de detalle del día: 5 columnas en lenguaje humano + resumen VDP."""
    corr_map = {c.nombre_norm: c for c in c4.correcciones}
    contable_map = {n: s for n, s in cont.sabores.items()} if cont else {}

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 46
    ws.sheet_view.showGridLines = False

    brd = _border()
    row = 1

    # ── Encabezado del día ────────────────────────────────────────
    ws.merge_cells(f'A{row}:E{row}')
    _cell(ws, row, 1, f'  DÍA {label}  —  Ventas de helados',
          font=_font(bold=True, color=C_WHITE, size=13),
          fill=_fill(C_HEADER_DARK),
          align=_align(h='left'))
    ws.row_dimensions[row].height = 30
    row += 1

    n_corr = len([c for c in c4.correcciones
                  if c.tipo_resolucion.value != 'FORZADO_H0'])
    n_forzado = len([c for c in c4.correcciones
                     if c.tipo_resolucion.value == 'FORZADO_H0'])
    n_sindat = sum(1 for sc in c3.sabores.values()
                   if sc.status in (StatusC3.SOLO_DIA, StatusC3.SOLO_NOCHE))

    partes = []
    if n_corr:
        partes.append(f'{n_corr} corrección{"es" if n_corr > 1 else ""} aplicada{"s" if n_corr > 1 else ""}')
    if n_forzado:
        partes.append(f'{n_forzado} estimada{"s" if n_forzado > 1 else ""} (revisar)')
    if n_sindat:
        partes.append(f'{n_sindat} sin datos de ambos turnos')
    resumen_txt = ' · '.join(partes) if partes else 'Sin correcciones — registros directos'

    ws.merge_cells(f'A{row}:E{row}')
    _cell(ws, row, 1, f'  {resumen_txt}',
          font=_font(italic=True, color='444444', size=9),
          fill=_fill('EEF4FF'),
          align=_align(h='left'))
    ws.row_dimensions[row].height = 18
    row += 2

    # ── Encabezados de columna ────────────────────────────────────
    headers = ['Sabor', 'Venta (g)', 'Ajuste (g)', 'Situación', 'Nota']
    hdr_fill = _fill(C_HEADER_MID)
    hdr_font = _font(bold=True, color=C_WHITE, size=10)
    for col, h in enumerate(headers, 1):
        al = _align(h='left' if col in (1, 5) else 'center')
        _cell(ws, row, col, h, hdr_font, hdr_fill, al, brd)
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Ordenar sabores: correcciones primero, luego por volumen ──
    def _sort_key(item):
        nombre, sc = item
        corr = corr_map.get(nombre)
        raw = contable_map[nombre].venta_raw if nombre in contable_map else 0
        sit, _, _, _ = _clasificar_fila(sc, corr, raw)
        return (_SITUACION_PRIORITY.get(sit, 9), nombre)

    sabores_sorted = sorted(c3.sabores.items(), key=_sort_key)

    # ── Filas de sabores ──────────────────────────────────────────
    for nombre, sc in sabores_sorted:
        corr = corr_map.get(nombre)
        raw = contable_map[nombre].venta_raw if nombre in contable_map else 0
        sit, color, delta, vf = _clasificar_fila(sc, corr, raw)
        nota = _nota_legible(sc, corr)
        row_fill = _fill(color)

        vals = [nombre, vf, delta, sit, nota]
        aligns = ['left', 'right', 'center', 'left', 'left']
        bolds  = [True, True, False, False, False]

        for col, (val, h, bold) in enumerate(zip(vals, aligns, bolds), 1):
            fmt = '#,##0' if col in (2, 3) and isinstance(val, (int, float)) else None
            _cell(ws, row, col, val,
                  _font(bold=bold, size=10),
                  row_fill,
                  _align(h=h, wrap=(col == 5)),
                  brd,
                  number_format=fmt)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Separador ────────────────────────────────────────────────
    row += 1

    # ── Subtotal helados ─────────────────────────────────────────
    ws.merge_cells(f'A{row}:C{row}')
    _cell(ws, row, 1, 'Subtotal ventas de helados',
          font=_font(bold=True, size=10),
          fill=_fill(C_TOTAL_ROW),
          align=_align(h='left', indent=1),
          border=_border(color='AAAAAA'))
    _cell(ws, row, 2)   # merged
    _cell(ws, row, 3)   # merged

    _cell(ws, row, 4, resultado.venta_refinado,
          font=_font(bold=True, size=10),
          fill=_fill(C_TOTAL_ROW),
          align=_align(h='right'),
          border=_border(color='AAAAAA'),
          number_format='#,##0')
    ws.merge_cells(f'D{row}:E{row}')
    ws.row_dimensions[row].height = 20
    row += 1

    # ── VDP ──────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:C{row}')
    _cell(ws, row, 1, '+ Ventas de postres (VDP)',
          font=_font(bold=True, size=10),
          fill=_fill(C_VDP_ROW),
          align=_align(h='left', indent=1),
          border=_border(color='AAAAAA'))
    _cell(ws, row, 4, resultado.vdp,
          font=_font(bold=True, size=10),
          fill=_fill(C_VDP_ROW),
          align=_align(h='right'),
          border=_border(color='AAAAAA'),
          number_format='#,##0')
    ws.merge_cells(f'D{row}:E{row}')
    ws.row_dimensions[row].height = 20
    row += 1

    # ── Aperturas ────────────────────────────────────────────────
    if resultado.n_latas > 0:
        texto_tapas = f'− Descuento aperturas ({resultado.n_latas} lata{"s" if resultado.n_latas > 1 else ""} × 280 g)'
        ws.merge_cells(f'A{row}:C{row}')
        _cell(ws, row, 1, texto_tapas,
              font=_font(bold=True, size=10),
              fill=_fill(C_TAPAS_ROW),
              align=_align(h='left', indent=1),
              border=_border(color='AAAAAA'))
        _cell(ws, row, 4, -resultado.lid_discount,
              font=_font(bold=True, size=10),
              fill=_fill(C_TAPAS_ROW),
              align=_align(h='right'),
              border=_border(color='AAAAAA'),
              number_format='#,##0')
        ws.merge_cells(f'D{row}:E{row}')
        ws.row_dimensions[row].height = 20
        row += 1

    # ── TOTAL DEL DÍA ────────────────────────────────────────────
    ws.merge_cells(f'A{row}:C{row}')
    _cell(ws, row, 1, 'TOTAL DEL DÍA',
          font=_font(bold=True, color=C_WHITE, size=12),
          fill=_fill(C_GRAND_TOTAL),
          align=_align(h='left', indent=1),
          border=_border(style='medium', color='1F4E79'))
    _cell(ws, row, 4, resultado.total_refinado,
          font=_font(bold=True, color=C_WHITE, size=12),
          fill=_fill(C_GRAND_TOTAL),
          align=_align(h='right'),
          border=_border(style='medium', color='1F4E79'),
          number_format='#,##0')
    ws.merge_cells(f'D{row}:E{row}')
    ws.row_dimensions[row].height = 26


# ===================================================================
# HOJA DETALLE TÉCNICO (trazabilidad)
# ===================================================================

def _write_detalle_tecnico(ws, resultados):
    """Trazabilidad completa para análisis avanzado."""
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 55
    ws.sheet_view.showGridLines = False

    hdr_font  = _font(bold=True, color=C_WHITE, size=9)
    hdr_fill  = _fill(C_HEADER_DARK)
    brd       = _border()

    headers = ['Día', 'Sabor', 'Raw (g)', 'Ajuste (g)', 'Final (g)',
               'Origen', 'Tipo', 'Banda', 'Conf.', 'Motivo técnico']
    for col, h in enumerate(headers, 1):
        _cell(ws, 1, col, h, hdr_font, hdr_fill, _align(wrap=True), brd)
    ws.row_dimensions[1].height = 22

    BANDA_FILLS = {
        Banda.CONFIRMADO: _fill('E8F5E9'),
        Banda.FORZADO:    _fill('FFF8E1'),
        Banda.ESTIMADO:   _fill('FFF8E1'),
    }

    row = 2
    for datos, cont, c3, c4, resultado in resultados:
        dia = f'D{resultado.dia_label}'

        # C3 prototipos
        for nombre in sorted(c3.sabores.keys()):
            sc = c3.sabores[nombre]
            if sc.prototipo:
                p = sc.prototipo
                rfill = _fill('E3F2FD')
                vals = [dia, nombre, sc.contable.venta_raw, p.delta, p.venta_corregida,
                        'C3', p.codigo, 'PF', f'{p.confianza:.0%}', p.descripcion]
                for col, val in enumerate(vals, 1):
                    fmt = '#,##0' if col in (3, 4, 5) else None
                    _cell(ws, row, col, val, _font(size=9), rfill,
                          _align(h='left' if col in (2, 10) else 'center', wrap=(col == 10)),
                          brd, number_format=fmt)
                row += 1

        # C4 correcciones
        for corr in sorted(c4.correcciones, key=lambda c: c.nombre_norm):
            rfill = BANDA_FILLS.get(corr.banda, _fill(C_WHITE))
            vals = [dia, corr.nombre_norm, corr.venta_raw, corr.delta, corr.venta_corregida,
                    'C4', corr.tipo_resolucion.value, corr.banda.value,
                    f'{corr.confianza:.0%}', corr.motivo]
            for col, val in enumerate(vals, 1):
                fmt = '#,##0' if col in (3, 4, 5) else None
                _cell(ws, row, col, val, _font(size=9), rfill,
                      _align(h='left' if col in (2, 10) else 'center', wrap=(col == 10)),
                      brd, number_format=fmt)
            row += 1


# ===================================================================
# FUNCIÓN PRINCIPAL
# ===================================================================

def exportar_multi(resultados, path_output: str):
    """
    Genera el Excel unificado.

    Args:
        resultados: lista de (datos, contabilidad, clasificacion, c4, resultado)
        path_output: ruta de salida
    """
    wb = openpyxl.Workbook()

    # Mapeo dia_label → nombre de hoja Excel
    sheet_names = {}
    for datos, cont, c3, c4, resultado in resultados:
        sheet_names[resultado.dia_label] = f'Día {resultado.dia_label}'

    # ── Hoja 1: Guía ─────────────────────────────────────────────
    ws_guia = wb.active
    ws_guia.title = 'Guía'
    _write_guia(ws_guia)

    # ── Hoja 2: Resumen ───────────────────────────────────────────
    ws_resumen = wb.create_sheet('Resumen')
    _write_resumen(ws_resumen, resultados, sheet_names)

    # ── Una hoja por día ──────────────────────────────────────────
    for datos, cont, c3, c4, resultado in resultados:
        nombre_hoja = sheet_names[resultado.dia_label]
        ws = wb.create_sheet(nombre_hoja)
        _write_dia(ws, resultado.dia_label, c3, c4, resultado, cont=cont)

    # ── Hoja final: Detalle técnico ───────────────────────────────
    ws_det = wb.create_sheet('Detalle técnico')
    _write_detalle_tecnico(ws_det, resultados)

    wb.save(path_output)
    print(f'Exportado: {path_output}')
