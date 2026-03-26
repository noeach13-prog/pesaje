"""
exporter.py -- Genera el Excel de salida v3 con bandas y trazabilidad.

Hojas:
  1. Ventas: una fila por sabor con raw, corregida, delta, banda
  2. Trazabilidad: cada correccion con cadena de evidencia
  3. Resumen: totales por banda, VDP, latas
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .modelos import (
    ResultadoC3, SaborClasificado, StatusC3,
    Correccion, Banda, TipoJustificacion,
    ResultadoDia,
)
from typing import List, Dict, Optional


# -- Colores --
C_HEADER = "1F4E79"
C_WHITE = "FFFFFF"
C_ALT_ROW = "F2F7FC"
C_CONFIRMADO = "E2EFDA"  # verde claro
C_FORZADO = "FFF2CC"     # amarillo claro
C_ESTIMADO = "FCE4EC"    # rosa claro
C_NEGATIVE = "FFE0B2"    # naranja
C_ENGINE = "E8F5E9"      # verde muy claro
C_PF = "E3F2FD"          # azul claro (prototipo)
C_LIMPIO = C_WHITE
C_SOLO = "F5F5F5"        # gris

BANDA_COLORS = {
    Banda.CONFIRMADO: C_CONFIRMADO,
    Banda.FORZADO: C_FORZADO,
    Banda.ESTIMADO: C_ESTIMADO,
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Arial")


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_border() -> Border:
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def _set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


# ===================================================================
# HOJA 1: Ventas por sabor
# ===================================================================

def _write_ventas(ws, clasificacion: ResultadoC3, correcciones: List[Correccion],
                  resultado: ResultadoDia):
    corr_map = {c.nombre_norm: c for c in correcciones}
    hdr_font = _font(bold=True, color=C_WHITE, size=10)
    hdr_fill = _fill(C_HEADER)
    hdr_align = _align(wrap=True)
    border = _thin_border()

    # Headers
    headers = ['Sabor', 'Status', 'Venta Raw', 'Delta', 'Venta Final',
               'Banda', 'Tipo', 'Confianza', 'Motivo']
    for col, h in enumerate(headers, 1):
        _set_cell(ws, 1, col, h, hdr_font, hdr_fill, hdr_align, border)

    # Ancho de columnas
    widths = [22, 12, 12, 10, 12, 14, 6, 10, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Datos: ordenar por status (SENAL/COMPUESTO primero, luego ENGINE, luego LIMPIO)
    order = {StatusC3.COMPUESTO: 0, StatusC3.SENAL: 1, StatusC3.ESCALAR: 2,
             StatusC3.ENGINE: 3, StatusC3.LIMPIO: 4,
             StatusC3.SOLO_DIA: 5, StatusC3.SOLO_NOCHE: 6, StatusC3.OBSERVACION: 7}
    sabores_sorted = sorted(clasificacion.sabores.items(),
                            key=lambda x: (order.get(x[1].status, 99), x[0]))

    row = 2
    for nombre, sc in sabores_sorted:
        corr = corr_map.get(nombre)
        raw = sc.contable.venta_raw
        has_proto = sc.prototipo is not None

        if corr:
            vf = corr.venta_corregida
            delta = corr.delta
            banda = corr.banda.value
            tipo = corr.tipo_justificacion.value
            conf = f'{corr.confianza:.0%}'
            motivo = corr.motivo
            row_fill = _fill(BANDA_COLORS.get(corr.banda, C_WHITE))
        elif has_proto:
            vf = sc.prototipo.venta_corregida
            delta = sc.prototipo.delta
            banda = 'PF'
            tipo = sc.prototipo.codigo
            conf = f'{sc.prototipo.confianza:.0%}'
            motivo = sc.prototipo.descripcion
            row_fill = _fill(C_PF)
        elif sc.venta_final_c3 is not None:
            vf = sc.venta_final_c3
            delta = 0
            banda = '-'
            tipo = '-'
            conf = '-'
            motivo = ''
            if sc.status == StatusC3.ENGINE:
                row_fill = _fill(C_ENGINE)
            elif sc.status in (StatusC3.SOLO_DIA, StatusC3.SOLO_NOCHE):
                row_fill = _fill(C_SOLO)
            else:
                row_fill = _fill(C_LIMPIO if row % 2 == 0 else C_ALT_ROW)
        elif sc.status in (StatusC3.SOLO_DIA, StatusC3.SOLO_NOCHE):
            # No calculable: solo un turno presente
            vf = 0  # display como 0 en Excel
            delta = 0
            banda = '-'
            tipo = '-'
            conf = '-'
            motivo = 'Solo un turno'
            row_fill = _fill(C_SOLO)
        else:
            # Sin resolver (H0)
            vf = raw
            delta = 0
            banda = 'H0'
            tipo = '-'
            conf = '-'
            motivo = 'Sin resolver'
            row_fill = _fill(C_NEGATIVE)

        vals = [nombre, sc.status.value, raw, delta, vf, banda, tipo, conf, motivo]
        for col, val in enumerate(vals, 1):
            f = _font(bold=(col == 1))
            a = _align(h='left' if col in (1, 9) else 'center', wrap=(col == 9))
            _set_cell(ws, row, col, val, f, row_fill, a, border)

        row += 1

    # Fila total
    row += 1
    _set_cell(ws, row, 1, 'TOTAL', _font(bold=True), _fill(C_HEADER), _align(), border)
    _set_cell(ws, row, 3, resultado.venta_raw, _font(bold=True, color=C_WHITE), _fill(C_HEADER), _align(), border)
    _set_cell(ws, row, 5, resultado.venta_refinado, _font(bold=True, color=C_WHITE), _fill(C_HEADER), _align(), border)

    row += 1
    _set_cell(ws, row, 1, 'VDP', _font(bold=True))
    _set_cell(ws, row, 5, resultado.vdp, _font(bold=True))

    row += 1
    _set_cell(ws, row, 1, 'Latas', _font(bold=True))
    _set_cell(ws, row, 5, f'{resultado.n_latas} ({resultado.lid_discount}g)', _font(bold=True))

    return row


# ===================================================================
# HOJA 2: Trazabilidad
# ===================================================================

def _write_trazabilidad(ws, correcciones: List[Correccion], clasificacion: ResultadoC3):
    hdr_font = _font(bold=True, color=C_WHITE)
    hdr_fill = _fill(C_HEADER)
    border = _thin_border()

    headers = ['Sabor', 'Raw', 'Corregida', 'Delta', 'Tipo', 'Banda',
               'Confianza', 'Resolucion', 'Motivo', 'Grupo']
    for col, h in enumerate(headers, 1):
        _set_cell(ws, 1, col, h, hdr_font, hdr_fill, _align(wrap=True), border)

    widths = [22, 10, 10, 10, 6, 14, 10, 18, 70, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Prototipos de Capa 3
    row = 2
    for nombre, sc in sorted(clasificacion.sabores.items()):
        if sc.prototipo:
            p = sc.prototipo
            row_fill = _fill(C_PF)
            vals = [nombre, sc.contable.venta_raw, p.venta_corregida, p.delta,
                    p.codigo, 'PF', f'{p.confianza:.0%}', 'PROTOTIPO_C3',
                    p.descripcion, '']
            for col, val in enumerate(vals, 1):
                _set_cell(ws, row, col, val, _font(), row_fill, _align(h='left' if col >= 9 else 'center', wrap=(col >= 9)), border)
            row += 1

    # Correcciones de Capa 4
    for corr in sorted(correcciones, key=lambda c: c.nombre_norm):
        row_fill = _fill(BANDA_COLORS.get(corr.banda, C_WHITE))
        vals = [corr.nombre_norm, corr.venta_raw, corr.venta_corregida, corr.delta,
                corr.tipo_justificacion.value, corr.banda.value,
                f'{corr.confianza:.0%}', corr.tipo_resolucion.value,
                corr.motivo, corr.grupo_conjunto or '']
        for col, val in enumerate(vals, 1):
            _set_cell(ws, row, col, val, _font(), row_fill, _align(h='left' if col >= 9 else 'center', wrap=(col >= 9)), border)
        row += 1


# ===================================================================
# HOJA 3: Resumen
# ===================================================================

def _write_resumen(ws, resultado: ResultadoDia, correcciones: List[Correccion]):
    hdr_font = _font(bold=True, color=C_WHITE)
    hdr_fill = _fill(C_HEADER)
    border = _thin_border()
    bold = _font(bold=True)

    row = 1
    _set_cell(ws, row, 1, f'Dia {resultado.dia_label}', _font(bold=True, size=14), None, _align(h='left'))
    row += 2

    # Totales por banda
    labels = [
        ('Venta RAW', resultado.venta_raw, None),
        ('+ CONFIRMADO', resultado.venta_confirmado - resultado.venta_raw, C_CONFIRMADO),
        ('+ FORZADO', resultado.venta_operativo - resultado.venta_confirmado, C_FORZADO),
        ('+ ESTIMADO', resultado.venta_refinado - resultado.venta_operativo, C_ESTIMADO),
        ('= Venta Refinada', resultado.venta_refinado, C_HEADER),
        ('', '', None),
        ('VDP', resultado.vdp, None),
        ('Latas', f'{resultado.n_latas} ({resultado.lid_discount}g)', None),
        ('', '', None),
        ('TOTAL OPERATIVO', resultado.total_operativo, C_HEADER),
    ]

    for label, val, color in labels:
        if not label:
            row += 1
            continue
        _set_cell(ws, row, 1, label, bold, _fill(color) if color else None, _align(h='left'), border)
        f = _font(bold=True, color=C_WHITE if color == C_HEADER else '000000')
        _set_cell(ws, row, 2, val, f, _fill(color) if color else None, _align(), border)
        row += 1

    # Stats
    row += 2
    _set_cell(ws, row, 1, 'Estadisticas', _font(bold=True, size=12))
    row += 1
    stats = [
        ('LIMPIO', resultado.n_limpio),
        ('ENGINE', resultado.n_engine),
        ('Escalados', resultado.n_escalado),
        ('Solo DIA/NOCHE', resultado.n_solo_dia),
        ('Correcciones C4', len(correcciones)),
    ]
    for label, val in stats:
        _set_cell(ws, row, 1, label, _font(), None, _align(h='left'), border)
        _set_cell(ws, row, 2, val, _font(), None, _align(), border)
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18


# ===================================================================
# FUNCION PRINCIPAL
# ===================================================================

def exportar(resultado: ResultadoDia, clasificacion: ResultadoC3,
             correcciones: List[Correccion], path_output: str):
    """Genera el Excel de salida v3."""
    wb = openpyxl.Workbook()

    # Hoja 1: Ventas
    ws1 = wb.active
    ws1.title = 'Ventas'
    _write_ventas(ws1, clasificacion, correcciones, resultado)

    # Hoja 2: Trazabilidad
    ws2 = wb.create_sheet('Trazabilidad')
    _write_trazabilidad(ws2, correcciones, clasificacion)

    # Hoja 3: Resumen
    ws3 = wb.create_sheet('Resumen')
    _write_resumen(ws3, resultado, correcciones)

    wb.save(path_output)
    return path_output
