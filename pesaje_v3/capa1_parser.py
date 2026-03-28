"""
capa1_parser.py — Wrapper sobre parser.py existente.
Lee el workbook y entrega datos crudos sin interpretación.
"""
import sys
import os
import re

# Agregar directorio padre al path para importar parser existente
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from parser import load_shifts, normalize_name
from models import ShiftData, FlavorShiftData

from .modelos import SaborCrudo, TurnoCrudo, DatosDia

import openpyxl


def _recuperar_ab_cero(path_excel: str, nombre_hoja: str, turno: 'TurnoCrudo'):
    """
    Fix R1.4: el parser v1 filtra sabores donde ab=0 y no hay cerradas/entrantes.
    Ej: CHOCOLATE DUBAI con ab=0 explicito.
    Re-leemos el Excel para recuperar estos sabores.
    """
    try:
        wb = openpyxl.load_workbook(path_excel, data_only=True)
        ws = wb[nombre_hoja]
    except (KeyError, FileNotFoundError):
        return

    for r in range(2, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        if a_val is None:
            continue
        if isinstance(a_val, str) and normalize_name(a_val).startswith('POSTRES'):
            break

        norm = normalize_name(a_val)
        if not norm or norm in turno.sabores:
            continue

        # Este sabor no fue incluido por el parser. Verificar si tiene ab=0 explicito.
        b_val = ws.cell(r, 2).value
        if b_val is not None and (b_val == 0 or b_val == 0.0):
            turno.sabores[norm] = SaborCrudo(
                nombre=str(a_val).strip(),
                nombre_norm=norm,
                abierta=0,
                celiaca=None,
                cerradas=[],
                entrantes=[],
            )


_PESO_MAX_FISICO = 7900  # gramos — T11: superar esto es imposible fisicamente


def _peso_valido(v) -> bool:
    """Devuelve True si el valor es un peso fisicamente posible."""
    return isinstance(v, (int, float)) and 0 < v <= _PESO_MAX_FISICO


def _shift_to_turno(shift: ShiftData) -> TurnoCrudo:
    """Convierte ShiftData (v1) a TurnoCrudo (v3)."""
    turno = TurnoCrudo(
        nombre_hoja=shift.name,
        indice=shift.index,
        vdp_textos=[s.raw_text for s in shift.ventas_sin_peso],
    )
    for norm_name, fsd in shift.flavors.items():
        turno.sabores[norm_name] = SaborCrudo(
            nombre=fsd.name,
            nombre_norm=norm_name,
            abierta=int(fsd.abierta) if fsd.abierta and _peso_valido(fsd.abierta) else None,
            celiaca=int(fsd.celiaca) if fsd.celiaca and _peso_valido(fsd.celiaca) else None,
            cerradas=[int(c) for c in fsd.cerradas if _peso_valido(c)],
            entrantes=[int(e) for e in fsd.entrantes if _peso_valido(e)],
        )

    # Fix R1.4: incluir sabores con ab=0 que el parser v1 filtra
    # El parser descarta sabores donde todo es 0/vacio, pero ab=0
    # es un dato explicito que Capa 1 debe preservar.
    # Tambien incluir sabores de la seccion observaciones.
    for obs in shift.observaciones:
        norm = normalize_name(obs.flavor)
        if norm and norm not in turno.sabores:
            ab = int(obs.abierta) if obs.abierta else None
            cerr = [int(obs.cerrada)] if obs.cerrada else []
            ent = [int(obs.entrante)] if obs.entrante else []
            if ab is not None or cerr or ent:
                turno.sabores[norm] = SaborCrudo(
                    nombre=obs.flavor,
                    nombre_norm=norm,
                    abierta=ab,
                    celiaca=None,
                    cerradas=cerr,
                    entrantes=ent,
                )

    turno.es_vacio = len(turno.sabores) == 0
    return turno


def _extraer_dia_label(nombre_hoja: str) -> str:
    """Extrae el número de día del nombre de hoja. Ej: 'Jueves 5 (DIA)' → '5'."""
    m = re.search(r'(\d+)', nombre_hoja)
    return m.group(1) if m else nombre_hoja


def _es_turno_dia(nombre: str) -> bool:
    return '(DIA)' in nombre.upper() or '(DÍA)' in nombre.upper() or 'DIA' in nombre.upper().split('(')[-1]


def _es_turno_noche(nombre: str) -> bool:
    return 'NOCHE' in nombre.upper()


_DIA_NAME_PATTERN = re.compile(
    r'(lunes|martes|mi[eé]rcoles|jueves|viernes|s[aá]bado|domingo)\s+\d+',
    re.IGNORECASE,
)


def _detectar_modo_workbook(shifts) -> str:
    """
    Detecta si el workbook usa pares DIA/NOCHE o turnos unicos.
    Retorna 'DIA_NOCHE' o 'TURNO_UNICO'.
    Aborta con ValueError si el formato es ambiguo o desconocido.
    """
    nombres = [s.name for s in shifts]

    tiene_dia = any(_es_turno_dia(n) for n in nombres)
    tiene_noche = any(_es_turno_noche(n) for n in nombres)

    if tiene_dia or tiene_noche:
        if tiene_dia and tiene_noche:
            return 'DIA_NOCHE'
        presente = 'DIA' if tiene_dia else 'NOCHE'
        faltante = 'NOCHE' if tiene_dia else 'DIA'
        raise ValueError(
            f"Workbook ambiguo: tiene hojas con {presente} pero ninguna con {faltante}. "
            "No se puede determinar el modo (DIA_NOCHE o TURNO_UNICO)."
        )

    # Sin marcadores DIA/NOCHE → esperar TURNO_UNICO.
    # Validacion secundaria: al menos una hoja debe tener nombre de dia plausible.
    plausibles = sum(1 for n in nombres if _DIA_NAME_PATTERN.search(n))
    if plausibles == 0:
        raise ValueError(
            "No se detectaron hojas DIA/NOCHE ni hojas con nombre de dia "
            "(ej: 'Domingo 1', 'Lunes 2'). Formato de workbook desconocido."
        )

    return 'TURNO_UNICO'


def cargar_dia(path_excel: str, dia_num: int, turnos_contexto: int = 3) -> DatosDia:
    """
    Carga un dia especifico del workbook.

    Detecta automaticamente el modo:
    - DIA_NOCHE: busca hojas con (DIA) y (NOCHE) para el mismo dia
    - TURNO_UNICO: trata hoja dia_num como turno A y dia_num+1 como turno B

    Args:
        path_excel: ruta al archivo Excel
        dia_num: numero del dia (ej: 5, 25, 28)
        turnos_contexto: cuantos turnos adyacentes cargar para Capa 4

    Returns:
        DatosDia con turno_dia, turno_noche y contexto
    """
    shifts = load_shifts(path_excel)
    modo = _detectar_modo_workbook(shifts)

    if modo == 'DIA_NOCHE':
        return _cargar_dia_dia_noche(path_excel, shifts, dia_num, turnos_contexto)
    else:
        return _cargar_dia_turno_unico(path_excel, shifts, dia_num, turnos_contexto)


def _cargar_dia_dia_noche(path_excel, shifts, dia_num, turnos_contexto):
    """Modo DIA/NOCHE: busca par (DIA, NOCHE) para el mismo dia."""
    dia_str = str(dia_num)

    turno_dia = None
    turno_noche = None
    idx_dia = -1
    idx_noche = -1

    for i, shift in enumerate(shifts):
        label = _extraer_dia_label(shift.name)
        if label == dia_str:
            turno = _shift_to_turno(shift)
            if _es_turno_dia(shift.name):
                turno_dia = turno
                idx_dia = i
            elif _es_turno_noche(shift.name):
                turno_noche = turno
                idx_noche = i

    if turno_dia is None:
        raise ValueError(f"No se encontro turno DIA para dia {dia_num}")
    if turno_noche is None:
        raise ValueError(f"No se encontro turno NOCHE para dia {dia_num}")

    # Fix R1.4: recuperar sabores con ab=0 que el parser v1 descarta
    _recuperar_ab_cero(path_excel, turno_dia.nombre_hoja, turno_dia)
    _recuperar_ab_cero(path_excel, turno_noche.nombre_hoja, turno_noche)

    # Contexto
    idx_min = max(0, min(idx_dia, idx_noche) - turnos_contexto)
    idx_max = min(len(shifts) - 1, max(idx_dia, idx_noche) + turnos_contexto)

    contexto = []
    for i in range(idx_min, idx_max + 1):
        if i == idx_dia or i == idx_noche:
            continue
        contexto.append(_shift_to_turno(shifts[i]))

    return DatosDia(
        dia_label=dia_str,
        turno_dia=turno_dia,
        turno_noche=turno_noche,
        contexto=contexto,
    )


def _cargar_dia_turno_unico(path_excel, shifts, dia_num, turnos_contexto):
    """
    Modo turno unico: hoja dia_num es turno A (referencia anterior),
    el siguiente dia con datos es turno B (dia cuya venta se calcula).
    Maneja gaps: dias cerrados (es_vacio=True) se saltan para encontrar B.
    """
    dia_str = str(dia_num)

    # Encontrar turno A por numero de dia
    idx_a = -1
    for i, shift in enumerate(shifts):
        if _extraer_dia_label(shift.name) == dia_str:
            idx_a = i
            break

    if idx_a == -1:
        raise ValueError(f"No se encontro hoja para dia {dia_num}")

    turno_a = _shift_to_turno(shifts[idx_a])

    # Encontrar turno B: siguiente hoja con datos despues de A
    idx_b = -1
    turno_b = None
    for i in range(idx_a + 1, len(shifts)):
        t = _shift_to_turno(shifts[i])
        if not t.es_vacio:
            idx_b = i
            turno_b = t
            break

    if turno_b is None:
        raise ValueError(
            f"No se encontro hoja con datos despues de dia {dia_num} (turno B)"
        )

    # Fix R1.4
    _recuperar_ab_cero(path_excel, turno_a.nombre_hoja, turno_a)
    _recuperar_ab_cero(path_excel, turno_b.nombre_hoja, turno_b)

    # La venta pertenece al dia B
    label = _extraer_dia_label(turno_b.nombre_hoja)

    # Contexto
    idx_min = max(0, idx_a - turnos_contexto)
    idx_max = min(len(shifts) - 1, idx_b + turnos_contexto)

    contexto = []
    for i in range(idx_min, idx_max + 1):
        if i == idx_a or i == idx_b:
            continue
        t = _shift_to_turno(shifts[i])
        if not t.es_vacio:
            contexto.append(t)

    return DatosDia(
        dia_label=label,
        turno_dia=turno_a,    # turno A = snapshot anterior (referencia)
        turno_noche=turno_b,  # turno B = snapshot actual (venta calculada aqui)
        contexto=contexto,
    )


def cargar_todos_los_dias(path_excel: str) -> list:
    """
    Carga todos los dias del workbook como lista de DatosDia.
    Detecta automaticamente el modo (DIA/NOCHE o turno unico).
    """
    shifts = load_shifts(path_excel)
    modo = _detectar_modo_workbook(shifts)
    turnos = [_shift_to_turno(s) for s in shifts]

    if modo == 'DIA_NOCHE':
        return _todos_los_dias_dia_noche(path_excel, turnos)
    else:
        return _todos_los_dias_turno_unico(path_excel, turnos)


def _todos_los_dias_dia_noche(path_excel, turnos):
    """Modo DIA/NOCHE: agrupar por numero de dia, buscar par."""
    dias_dict = {}
    for i, turno in enumerate(turnos):
        if turno.es_vacio:
            continue
        label = _extraer_dia_label(turno.nombre_hoja)
        if label not in dias_dict:
            dias_dict[label] = {'turnos': [], 'indices': []}
        dias_dict[label]['turnos'].append(turno)
        dias_dict[label]['indices'].append(i)

    resultado = []
    for label, info in dias_dict.items():
        dia_turno = None
        noche_turno = None
        for t in info['turnos']:
            if _es_turno_dia(t.nombre_hoja):
                dia_turno = t
            elif _es_turno_noche(t.nombre_hoja):
                noche_turno = t

        if dia_turno and noche_turno:
            idx_min = max(0, min(info['indices']) - 3)
            idx_max = min(len(turnos) - 1, max(info['indices']) + 3)
            contexto = [turnos[i] for i in range(idx_min, idx_max + 1)
                        if i not in info['indices'] and not turnos[i].es_vacio]
            resultado.append(DatosDia(
                dia_label=label,
                turno_dia=dia_turno,
                turno_noche=noche_turno,
                contexto=contexto,
            ))

    return resultado


def _todos_los_dias_turno_unico(path_excel, turnos):
    """Modo turno unico: pares consecutivos (dia N, dia N+1)."""
    # Filtrar turnos vacios y hojas STOCK
    validos = [(i, t) for i, t in enumerate(turnos) if not t.es_vacio]

    resultado = []
    for j in range(len(validos) - 1):
        idx_a, turno_a = validos[j]
        idx_b, turno_b = validos[j + 1]
        label = _extraer_dia_label(turno_b.nombre_hoja)  # venta pertenece al dia B

        # Fix R1.4
        _recuperar_ab_cero(path_excel, turno_a.nombre_hoja, turno_a)
        _recuperar_ab_cero(path_excel, turno_b.nombre_hoja, turno_b)

        # Contexto
        idx_min = max(0, idx_a - 3)
        idx_max = min(len(turnos) - 1, idx_b + 3)
        contexto = [turnos[i] for i in range(idx_min, idx_max + 1)
                     if i != idx_a and i != idx_b and not turnos[i].es_vacio]

        resultado.append(DatosDia(
            dia_label=label,
            turno_dia=turno_a,
            turno_noche=turno_b,
            contexto=contexto,
        ))

    return resultado
