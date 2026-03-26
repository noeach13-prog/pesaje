"""Debug 2: verificar sabores faltantes directamente en el Excel"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl

EXCEL = r"C:\Users\EliteBook\Downloads\Febrero San Martin 2026 (1).xlsx"
wb = openpyxl.load_workbook(EXCEL, data_only=True)

# Buscar KITKAT en DIA
for name in wb.sheetnames:
    if 'Jueves 5' in name and 'DIA' in name.upper():
        ws = wb[name]
        print(f"=== {name} ===")
        for r in range(2, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a and isinstance(a, str) and 'KIT' in a.upper():
                b = ws.cell(r, 2).value
                d = ws.cell(r, 4).value
                print(f"  Row {r}: A={a!r}  B={b!r}  D={d!r}")
            if a and isinstance(a, str) and 'COCO' in a.upper():
                b = ws.cell(r, 2).value
                print(f"  Row {r}: A={a!r}  B={b!r}")
            if a and isinstance(a, str) and 'DUBAI' in a.upper():
                b = ws.cell(r, 2).value
                print(f"  Row {r}: A={a!r}  B={b!r}")
            if a and isinstance(a, str) and 'KIYKAT' in a.upper():
                b = ws.cell(r, 2).value
                d = ws.cell(r, 4).value
                print(f"  Row {r}: A={a!r}  B={b!r}  D={d!r}")

# Buscar en NOCHE tambien
for name in wb.sheetnames:
    if 'Jueves 5' in name and 'NOCHE' in name.upper():
        ws = wb[name]
        print(f"\n=== {name} ===")
        for r in range(2, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a and isinstance(a, str) and ('KIT' in a.upper() or 'KIY' in a.upper()):
                b = ws.cell(r, 2).value
                d = ws.cell(r, 4).value
                print(f"  Row {r}: A={a!r}  B={b!r}  D={d!r}")
            if a and isinstance(a, str) and 'COCO' in a.upper():
                b = ws.cell(r, 2).value
                print(f"  Row {r}: A={a!r}  B={b!r}")
            if a and isinstance(a, str) and 'DUBAI' in a.upper():
                b = ws.cell(r, 2).value
                print(f"  Row {r}: A={a!r}  B={b!r}")

# Check normalize_name for KIYKAT
from parser import normalize_name
print(f"\nnormalize_name('KITKAT') = {normalize_name('KITKAT')!r}")
print(f"normalize_name('KIYKAT') = {normalize_name('KIYKAT')!r}")
print(f"normalize_name('Chocolate Dubai') = {normalize_name('Chocolate Dubai')!r}")
print(f"normalize_name('COCO') = {normalize_name('COCO')!r}")

# Check what the manual d5 had for these
print("\n=== Valores manuales esperados ===")
print("KITKAT DIA: ab=4630 cerr=[6400]")
print("KITKAT NOCHE: ab=4400 cerr=[]  (listed as KIYKAT)")
print("COCO DIA: ab=4505 cerr=[] (from POSTRES notes)")
print("COCO NOCHE: ab=4080 cerr=[]")
print("CHOC DUBAI DIA: ab=0 cerr=[] (empty)")
print("CHOC DUBAI NOCHE: empty")

# Calculate the deltas
print("\n=== Impacto en RAW ===")
# KITKAT: manual tiene DIA=(4630+6400)=11030, NOCHE=4400 -> raw=11030-4400-280=6350
# Pipeline tiene KITKAT SOLO_NOCHE=0 (no cuenta)
print("KITKAT: manual raw=6350, pipeline raw=0 (SOLO_NOCHE). Delta = +6350")

# COCO: manual DIA=4505 NOCHE=4080 -> raw=425
# Pipeline COCO SOLO_NOCHE=0
print("COCO: manual raw=425, pipeline raw=0 (SOLO_NOCHE). Delta = +425")

# Total expected delta
print(f"\nDelta esperado: 6350 + 425 = 6775")
print(f"Delta real: 6775. MATCH!")
