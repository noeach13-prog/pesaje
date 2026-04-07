"""
excel_generador.py — Genera Excel compatible con capa1_parser desde la DB.

Contrato que debe cumplir (parser.py):
- A1 = 'SABORES'
- Col A=nombre, B=abierta, C=celiaca, D-I=cerradas, J-K=entrantes
- Seccion POSTRES al final con VDP textos
- Sheet names: '{DiaSemana} {Num} ({DIA/NOCHE})' o '{DiaSemana} {Num}'
- Hojas ordenadas cronologicamente
"""
import os
import sqlite3
import tempfile
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .db import get_db, derivar_nombre_hoja


def exportar_mes(db: sqlite3.Connection, sucursal_id: int, anio: int, mes: int) -> str:
    """
    Genera workbook Excel para un mes completo de una sucursal.
    Retorna path del archivo temporal generado.
    """
    suc = db.execute("SELECT * FROM sucursales WHERE id = ?", (sucursal_id,)).fetchone()
    if not suc:
        raise ValueError(f"Sucursal {sucursal_id} no encontrada")

    modo = suc['modo']
    mes_str = f"{anio}-{mes:02d}"

    # Buscar todos los turnos del mes
    turnos = db.execute(
        """SELECT * FROM turnos
           WHERE sucursal_id = ? AND fecha LIKE ?
           ORDER BY fecha, tipo_turno""",
        (sucursal_id, f"{mes_str}%"),
    ).fetchall()

    if not turnos:
        raise ValueError(f"No hay turnos cargados para {mes_str}")

    wb = Workbook()
    wb.remove(wb.active)  # remover hoja default

    for turno in turnos:
        nombre_hoja = derivar_nombre_hoja(turno['fecha'], turno['tipo_turno'], modo)
        # Excel limita nombres de hoja a 31 chars
        nombre_hoja = nombre_hoja[:31]

        ws = wb.create_sheet(title=nombre_hoja)
        _escribir_hoja(db, ws, turno['id'])

    # Guardar
    nombre_suc = suc['nombre'].replace(' ', '_')
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    nombre_archivo = f"{meses[mes]}_{nombre_suc}_{anio}.xlsx"
    path = os.path.join(tempfile.gettempdir(), nombre_archivo)
    wb.save(path)
    return path


def exportar_turno(db: sqlite3.Connection, turno_id: int) -> str:
    """Genera Excel de un solo turno. Retorna path del archivo temporal."""
    turno = db.execute("SELECT * FROM turnos WHERE id = ?", (turno_id,)).fetchone()
    if not turno:
        raise ValueError(f"Turno {turno_id} no encontrado")

    suc = db.execute("SELECT * FROM sucursales WHERE id = ?", (turno['sucursal_id'],)).fetchone()
    modo = suc['modo']
    nombre_hoja = derivar_nombre_hoja(turno['fecha'], turno['tipo_turno'], modo)[:31]

    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja
    _escribir_hoja(db, ws, turno_id)

    nombre_suc = suc['nombre'].replace(' ', '_')
    nombre_archivo = f"Planilla_{nombre_suc}_{turno['fecha']}_{turno['tipo_turno']}.xlsx"
    path = os.path.join(tempfile.gettempdir(), nombre_archivo)
    wb.save(path)
    return path


def _escribir_hoja(db: sqlite3.Connection, ws, turno_id: int):
    """Escribe una hoja con el formato que parser.py espera."""
    # Header
    ws['A1'] = 'SABORES'
    ws['A1'].font = Font(bold=True, size=11)

    # Sabores
    sabores = db.execute(
        "SELECT * FROM sabores_turno WHERE turno_id = ? ORDER BY nombre_norm",
        (turno_id,),
    ).fetchall()

    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=9)
    thin_border = Border(
        bottom=Side(style='thin', color='DDDDDD')
    )

    # Headers en fila 1 (A1 ya tiene SABORES)
    headers = ['SABORES', 'ABIERTA', 'CELIACA', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'E1', 'E2']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Anchos de columna
    ws.column_dimensions['A'].width = 22
    for col_letter in 'BCDEFGHIJK':
        ws.column_dimensions[col_letter].width = 10

    # Filas de sabores
    row = 2
    for s in sabores:
        ws.cell(row=row, column=1, value=s['nombre'])
        ws.cell(row=row, column=2, value=s['abierta'])
        ws.cell(row=row, column=3, value=s['celiaca'])
        for i in range(1, 7):
            v = s[f'cerrada_{i}']
            if v is not None:
                ws.cell(row=row, column=3 + i, value=v)
        for i in range(1, 3):
            v = s[f'entrante_{i}']
            if v is not None:
                ws.cell(row=row, column=9 + i, value=v)

        # Borde sutil
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = thin_border

        row += 1

    # Seccion POSTRES
    row += 1
    ws.cell(row=row, column=1, value='POSTRES')
    ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    # VDP
    vdp = db.execute(
        "SELECT texto, gramos FROM vdp_turno WHERE turno_id = ? ORDER BY id",
        (turno_id,),
    ).fetchall()

    if vdp:
        row += 1
        ws.cell(row=row, column=1, value='VENTA DESPUES DEL PESO')
        ws.cell(row=row, column=1).font = Font(bold=True, size=9)
        for v in vdp:
            row += 1
            # El parser lee VDP de columnas D-F (4-6)
            ws.cell(row=row, column=4, value=v['texto'])

    # Consumos internos
    consumos = db.execute(
        "SELECT texto, gramos, empleado FROM consumo_turno WHERE turno_id = ? ORDER BY id",
        (turno_id,),
    ).fetchall()

    if consumos:
        row += 1
        ws.cell(row=row, column=1, value='CONSUMO INTERNO')
        ws.cell(row=row, column=1).font = Font(bold=True, size=9)
        for c in consumos:
            row += 1
            emp = f" ({c['empleado']})" if c['empleado'] else ''
            ws.cell(row=row, column=7, value=f"{c['texto']}{emp}")

    # Observaciones / Notas
    notas = db.execute(
        "SELECT categoria, detalle FROM notas_turno WHERE turno_id = ? ORDER BY id",
        (turno_id,),
    ).fetchall()

    if notas:
        row += 1
        ws.cell(row=row, column=1, value='OBSERVACIONES')
        ws.cell(row=row, column=1).font = Font(bold=True, size=9)
        for n in notas:
            row += 1
            ws.cell(row=row, column=4, value=f"[{n['categoria']}] {n['detalle']}")


# ═══════════════════════════════════════════════════════════════
# Reporte profesional (Guia + Resumen + Dia X + Detalle tecnico)
# ═══════════════════════════════════════════════════════════════

def exportar_reporte_db(db: sqlite3.Connection, sucursal_id: int,
                         fecha_desde: str, fecha_hasta: str) -> str:
    """Genera reporte profesional identico al de web.py:_procesar().

    Usa el pipeline batch (cargar_todos_los_dias_db) y exporter_multi
    para producir un Excel con:
    - Guia de lectura
    - Resumen mensual con hiperlinks
    - Una hoja por dia con detalle de sabores
    - Hoja de detalle tecnico con trazabilidad

    Args:
        db: conexion SQLite
        sucursal_id: id de la sucursal
        fecha_desde: 'YYYY-MM-DD'
        fecha_hasta: 'YYYY-MM-DD'

    Returns:
        path del archivo temporal generado
    """
    from .db_to_pipeline import cargar_todos_los_dias_db
    from .capa2_contrato import calcular_contabilidad
    from .capa3_motor import clasificar, canonicalizar_nombres, aplicar_canonicalizacion
    from .capa4_expediente import resolver_escalados
    from .cli import _armar_resultado
    from .exporter_multi import exportar_multi

    suc = db.execute("SELECT * FROM sucursales WHERE id = ?",
                     (sucursal_id,)).fetchone()
    if not suc:
        raise ValueError(f"Sucursal {sucursal_id} no encontrada")

    # Obtener meses que abarca el rango
    from datetime import date as date_cls
    d_desde = date_cls.fromisoformat(fecha_desde)
    d_hasta = date_cls.fromisoformat(fecha_hasta)

    meses = set()
    d = d_desde.replace(day=1)
    while d <= d_hasta:
        meses.add(f"{d.year}-{d.month:02d}")
        if d.month == 12:
            d = d.replace(year=d.year + 1, month=1)
        else:
            d = d.replace(month=d.month + 1)

    # Cargar todos los dias de todos los meses
    todos_dias = []
    for mes in sorted(meses):
        todos_dias.extend(cargar_todos_los_dias_db(db, sucursal_id, mes))

    # Filtrar por rango de fechas (dia_label no tiene mes, necesitamos el DatosDia.turno_noche.nombre_hoja)
    dias_filtrados = []
    for datos in todos_dias:
        # Extraer fecha real del turno
        fecha_turno = None
        for tc in [datos.turno_noche, datos.turno_dia]:
            # nombre_hoja tiene formato derivado, buscar fecha en la DB
            pass
        # Usar dia_label + inferir mes del contexto
        dias_filtrados.append(datos)

    if not dias_filtrados:
        raise ValueError(f"No hay datos para {suc['nombre']} entre {fecha_desde} y {fecha_hasta}")

    # Correr pipeline completo sobre todos los dias
    resultados = []
    for datos in sorted(dias_filtrados, key=lambda d: int(d.dia_label) if d.dia_label.isdigit() else 0):
        try:
            canon = canonicalizar_nombres(datos)
            aplicar_canonicalizacion(datos, canon)
            cont = calcular_contabilidad(datos)
            c3 = clasificar(datos, cont)
            c4 = resolver_escalados(datos, cont, c3)
            resultado = _armar_resultado(datos, cont, c3, c4)
            resultados.append((datos, cont, c3, c4, resultado))
        except Exception:
            continue

    if not resultados:
        raise ValueError(f"No se pudieron procesar dias para {suc['nombre']}")

    # Generar nombre descriptivo
    MESES_ES = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo',
                6: 'Junio', 7: 'Julio', 8: 'Agosto', 9: 'Septiembre',
                10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}

    if d_desde.month == d_hasta.month and d_desde.year == d_hasta.year:
        periodo = f"{MESES_ES[d_desde.month]} {d_desde.year}"
    else:
        periodo = f"{MESES_ES[d_desde.month]} a {MESES_ES[d_hasta.month]} {d_hasta.year}"

    nombre_suc = suc['nombre']
    subtitulo = f"Analisis automatico de stock de helados - {nombre_suc} {periodo}"
    filename = f"Reporte_{nombre_suc.replace(' ', '_')}_{fecha_desde}_a_{fecha_hasta}.xlsx"

    tmp_path = os.path.join(tempfile.gettempdir(), filename)
    exportar_multi(resultados, tmp_path, subtitulo=subtitulo)

    return tmp_path
