"""Debug 3: que dice exactamente el Excel para KITKAT en DIA"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import openpyxl

EXCEL = r"C:\Users\EliteBook\Downloads\Febrero San Martin 2026 (1).xlsx"
wb = openpyxl.load_workbook(EXCEL, data_only=True)

for name in wb.sheetnames:
    if 'Jueves 5' in name and 'DIA' in name.upper():
        ws = wb[name]
        print(f"=== {name} ===")
        # Search ALL rows for anything with KIT
        for r in range(2, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a and isinstance(a, str) and 'KIT' in a.upper():
                vals = [ws.cell(r, c).value for c in range(1, 12)]
                print(f"  Row {r}: {vals}")

        # Also check observaciones section
        print("\n  --- Observaciones ---")
        in_obs = False
        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, 4).value
            if d and isinstance(d, str) and 'OBSERVACION' in d.upper():
                in_obs = True
                continue
            if in_obs:
                vals_d = ws.cell(r, 4).value
                if vals_d and isinstance(vals_d, str) and 'KIT' in vals_d.upper():
                    row_vals = [ws.cell(r, c).value for c in range(1, 12)]
                    print(f"  Row {r}: {row_vals}")

from parser import normalize_name
print(f"\nnormalize_name('KIT KAT') = {normalize_name('KIT KAT')!r}")
print(f"normalize_name('KITKAT') = {normalize_name('KITKAT')!r}")
print(f"normalize_name('KIYKAT') = {normalize_name('KIYKAT')!r}")
