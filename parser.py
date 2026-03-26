"""
parser.py — Reads the input Excel and returns a list of ShiftData objects.
"""
import re
import datetime
from typing import Optional
import openpyxl
from models import (ShiftData, FlavorShiftData, SaleEntry, ConsumoEntry, ObservacionEntry,
                     RawShift, RawFlavorObs, SlotValue, V2ShiftAnnotations)

# ── Constants ────────────────────────────────────────────────────────────────

EMPLOYEE_NAMES = {
    'samanta', 'ana', 'adry', 'adri', 'martin', 'carlos',
    'checho', 'sebastian', 'adriana', 'roco', 'roberto',
}

# text_to_grams pattern table: (regex, grams_per_unit)
# Equivalences (confirmed by operator):
#   cucurucho / vaso 4              = 220g
#   vaso 65 / vaso 3                = 180g
#   2 bochas / cucuruchon / mini    = 140g
#   vasito / vaso 2 / milkshake     = 140g
#   cono americano / 1 bocha / vaso 1 = 70g
#
# NOTE: "bochas" works as scoop count: N bochas = N × 70g.
#       "2 bochas" = 140g, "UNO DE 2 BOCHAS" = 140g.
#       "cucurucho" (220g) ≠ "cucuruchon" / "minicucuruchon" (140g).
_UNIT_PATTERNS = [
    # ── Kilos / fractions ──
    (r'(?<!/)(\d+(?:[.,]\d+)?)\s*(?:kilo|kilos|kg)\b',      1000),
    (r'(\d+(?:[.,]\d+)?)\s*1/2\s*(?:kilo|kg)',              500),
    (r'1/2\b',                                               500),
    (r'medio\b',                                             500),
    (r'(\d+(?:[.,]\d+)?)\s*1/4\b',                          250),
    (r'(?<!\d\s)1/4\b',                                      250),

    # ── 220g: cucurucho, vaso 4 ──
    (r'(\d+(?:[.,]\d+)?)\s*cucuruchos?\b',                   220),
    (r'(?<!\d\s)(?<!\d)\bcucuruchos?\b',                     220),
    (r'(\d+(?:[.,]\d+)?)\s*vasos?\s*4\b',                   220),
    (r'(?<!\d\s)\bvaso\s*4\b',                               220),

    # ── 180g: vaso 65, vasito 65, vaso 3 ──
    (r'(\d+(?:[.,]\d+)?)\s*(?:vasos?\s*65|vasitos?\s*65)\b', 180),
    (r'(?<!\d\s)(?:vasos?|vasitos?)\s*65\b',                 180),
    (r'(\d+(?:[.,]\d+)?)\s*vasos?\s*3\b',                   180),
    (r'(?<!\d\s)\bvaso\s*3\b',                               180),

    # ── 140g: cucuruchon / minicucuruchon, vasito, vaso 2, milkshake, blister ──
    (r'(\d+(?:[.,]\d+)?)\s*(?:cucuruchon(?:es)?|cucucruchon(?:es)?)\b', 140),
    (r'(?<!\d\s)\b(?:cucuruchon(?:es)?|cucucruchon(?:es)?)\b',          140),
    (r'(\d+(?:[.,]\d+)?)\s*(?:minicucuruchon(?:es)?|mini\s*cucuruchos?)\b', 140),
    (r'(?<!\d\s)\b(?:minicucuruchon(?:es)?|mini\s*cucuruchos?)\b',          140),
    (r'(\d+(?:[.,]\d+)?)\s*vasos?\s*2\b',                   140),
    (r'(?<!\d\s)\bvaso\s*2\b',                               140),
    (r'(\d+(?:[.,]\d+)?)\s*vasitos?(?!\s*\d)\b',            140),
    (r'(?<!\d\s)\bvasitos?(?!\s*\d)\b',                      140),
    (r'(\d+(?:[.,]\d+)?)\s*milkshakes?\b',                  140),
    (r'(?<!\d\s)\bmilkshakes?\b',                            140),
    (r'(\d+(?:[.,]\d+)?)\s*(?:bl\.\s*vasito|bl\s+vasito|blister\s+vasito)s?\b', 140),

    # ── 70g: cono americano, bocha (per scoop), vaso 1 ──
    (r'(\d+(?:[.,]\d+)?)\s*conos?\s+americanos?\b',         70),
    (r'(?<!\d\s)\bconos?\s+americanos?\b',                   70),
    (r'(\d+(?:[.,]\d+)?)\s*bochas?\b',                      70),
    (r'(?<!\d\s)\bbochas?\b',                                70),
    (r'(\d+(?:[.,]\d+)?)\s*vasos?\s*(?:1|uno)\b',           70),
    (r'(?<!\d\s)\bvaso\s*(?:1|uno)\b',                       70),

    # ── kilo standalone ──
    (r'(?<!\d\s)(?<!\d)\bkilos?\b',                          1000),

    # ── Zero-gram (skip) ──
    (r'(?:agua\b|gio\b)',                                    0),
]

# Compile patterns once
_COMPILED = [(re.compile(pat, re.IGNORECASE), g) for pat, g in _UNIT_PATTERNS]


def text_to_grams(text: str) -> float:
    """Convert a free-text ice-cream unit description to grams."""
    if not text or not isinstance(text, str):
        return 0.0
    text = str(text).strip()
    # Remove employee names and noise words
    for emp in EMPLOYEE_NAMES:
        text = re.sub(rf'\b{emp}\b', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\b(promo|uno|una|de|del|con|mini)\b', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\s+', ' ', text).strip()  # collapse double-spaces left by noise removal

    total = 0.0
    for pattern, grams_unit in _COMPILED:
        for m in pattern.finditer(text):
            try:
                qty = float(m.group(1).replace(',', '.')) if m.lastindex and m.group(1) else 1.0
            except (IndexError, AttributeError):
                qty = 1.0
            total += qty * grams_unit
    return total


# Known flavor name typos/variants → canonical form.
# Applied AFTER accent removal and uppercasing.
_NAME_ALIASES = {
    'TIRAMIZU': 'TIRAMISU',
}


def normalize_name(name) -> str:
    """Normalize a flavor name for dict key comparison."""
    if name is None:
        return ''
    s = str(name).strip().upper()
    # Remove accents for comparison
    accents = str.maketrans('ÁÉÍÓÚÜÑ', 'AEIOUUN')
    s = s.translate(accents)
    # Collapse extra spaces
    s = re.sub(r'\s+', ' ', s)
    # Apply known name aliases
    s = _NAME_ALIASES.get(s, s)
    return s


def _safe_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    if isinstance(val, datetime.datetime):
        return None  # Excel date misparse
    if isinstance(val, str):
        s = val.strip().replace(',', '.')
        if s in ('-', '', '0') :
            return 0.0 if s == '0' else None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _is_employee(text) -> Optional[str]:
    """Return normalized employee name if text starts with one, else None."""
    if not text or not isinstance(text, str):
        return None
    s = text.strip().lower()
    # Handle leading '('
    s = s.lstrip('( ')
    for emp in EMPLOYEE_NAMES:
        if s.startswith(emp):
            return emp
    return None


# ── Row helpers ───────────────────────────────────────────────────────────────

def _parse_ventas_text(cell_value) -> list:
    """Extract SaleEntry items from a cell that may contain multiple unit descriptions."""
    if cell_value is None:
        return []
    if isinstance(cell_value, datetime.datetime):
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    # Skip pure-numeric or employee-only entries
    if re.match(r'^\d+([.,]\d+)?$', text):
        return []
    grams = text_to_grams(text)
    if grams > 0 or any(kw in text.lower() for kw in ['vasito', 'kilo', 'cucurucho', 'cucuruchon', 'cono', 'vaso', 'bocha', 'milkshake', 'kg', '1/2', '1/4']):
        return [SaleEntry(raw_text=text, grams=grams)]
    return []


# ── Main parser ───────────────────────────────────────────────────────────────

def load_shifts(path: str) -> list:
    """Load all valid shifts from the workbook. Returns list[ShiftData]."""
    wb = openpyxl.load_workbook(path, data_only=True)
    shifts = []

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        if ws['A1'].value != 'SABORES':
            continue  # skip STOCK sheets and others

        shift = ShiftData(name=sheet_name.strip(), index=idx)
        _parse_sheet(ws, shift)
        shifts.append(shift)

    return shifts


def _parse_sheet(ws, shift: ShiftData):
    max_row = ws.max_row

    postres_row = None
    flavor_rows = []

    # ── Pass 1: find SABORES section and POSTRES marker ──────────────────────
    for r in range(2, max_row + 1):
        a_val = ws.cell(r, 1).value
        # Check for POSTRES
        if a_val is not None and isinstance(a_val, str) and normalize_name(a_val).startswith('POSTRES'):
            postres_row = r
            break
        # Skip totals row (A is None, rest are large numbers)
        if a_val is None:
            b_val = _safe_float(ws.cell(r, 2).value)
            if b_val is not None and b_val > 10000:
                continue  # totals row
            continue
        # Skip header row if repeated
        if normalize_name(a_val) == 'SABORES':
            continue
        flavor_rows.append(r)

    # ── Parse flavor rows ────────────────────────────────────────────────────
    for r in flavor_rows:
        name_raw = ws.cell(r, 1).value
        if name_raw is None:
            continue
        name_norm = normalize_name(name_raw)
        if not name_norm or name_norm in ('', 'SABORES', 'POSTRES'):
            continue

        abierta = _safe_float(ws.cell(r, 2).value) or 0.0
        celiaca = _safe_float(ws.cell(r, 3).value) or 0.0

        cerradas = []
        for c in range(4, 10):  # cols D-I (openpyxl 1-indexed)
            v = _safe_float(ws.cell(r, c).value)
            if v is not None and v > 0:
                cerradas.append(v)

        entrantes = []
        for c in range(10, 12):  # cols J-K
            v = _safe_float(ws.cell(r, c).value)
            if v is not None and v > 0:
                entrantes.append(v)

        # Only store if has any data
        if abierta > 0 or celiaca > 0 or cerradas or entrantes:
            shift.flavors[name_norm] = FlavorShiftData(
                name=name_raw.strip(),
                abierta=abierta,
                celiaca=celiaca,
                cerradas=cerradas,
                entrantes=entrantes,
            )

    # ── Parse POSTRES section ────────────────────────────────────────────────
    if postres_row:
        _parse_postres(ws, postres_row + 1, max_row, shift)


def _parse_postres(ws, start_row: int, end_row: int, shift: ShiftData):
    """Parse VENTAS, CONSUMO INTERNO, and OBSERVACIONES from the POSTRES block."""
    in_observaciones = False
    obs_col_map = {'abierta': 5, 'cerrada': 7, 'entrante': 8}  # 1-indexed

    for r in range(start_row, end_row + 1):
        # Read relevant cells
        col_a = ws.cell(r, 1).value   # POSTRES items list
        col_d = ws.cell(r, 4).value   # VENTA DESPUES DEL PESO
        col_e = ws.cell(r, 5).value
        col_f = ws.cell(r, 6).value
        col_g = ws.cell(r, 7).value   # CONSUMO INTERNO
        col_h = ws.cell(r, 8).value
        col_i = ws.cell(r, 9).value

        # Detect OBSERVACIONES section marker
        if col_d is not None and isinstance(col_d, str) and 'OBSERVACION' in col_d.upper():
            in_observaciones = True
            continue

        # Detect OBSERVACIONES sub-header (ABIERTO/ABIERTA in col_e)
        if in_observaciones and col_e is not None and isinstance(col_e, str) and 'ABIERTO' in col_e.upper():
            # Detect column mapping from this row
            for c_idx in range(4, 12):
                cell_val = ws.cell(r, c_idx).value
                if cell_val is not None and isinstance(cell_val, str):
                    up = cell_val.upper()
                    if 'ABIERTO' in up or 'ABIERTA' in up:
                        obs_col_map['abierta'] = c_idx
                    elif 'CERRADO' in up or 'CERRADA' in up:
                        obs_col_map['cerrada'] = c_idx
                    elif 'ENTRANTE' in up:
                        obs_col_map['entrante'] = c_idx
            continue

        if in_observaciones:
            # Col D should be the flavor name (or free-text note)
            if col_d is None or isinstance(col_d, (int, float)):
                # Check if it's totally empty
                if all(ws.cell(r, c).value is None for c in range(4, 10)):
                    continue
            name_obs = col_d
            if name_obs is None:
                continue
            if isinstance(name_obs, datetime.datetime):
                continue
            name_obs_str = str(name_obs).strip()
            if not name_obs_str:
                continue

            ab_val = _safe_float(ws.cell(r, obs_col_map['abierta']).value)
            ce_val = _safe_float(ws.cell(r, obs_col_map['cerrada']).value)
            en_val = _safe_float(ws.cell(r, obs_col_map.get('entrante', 8)).value)

            # Determine if stock observation or free text
            if ab_val is not None or ce_val is not None or en_val is not None:
                shift.observaciones.append(ObservacionEntry(
                    flavor=name_obs_str,
                    abierta=ab_val,
                    cerrada=ce_val,
                    entrante=en_val,
                ))
            elif name_obs_str.upper() not in {
                'DELIVERY', 'DIA', 'NOCHE', 'INGRESO', 'CIERRE',
                'HELADERIA EFECTIVO:', 'HELADERIA ON LINE:',
                'PEDIDOS YA EFECTIVO:', 'PEDIDOS YA ON LINE:',
                'VENTA ANTES DEL PESAJE', 'VENTA DESPUES DEL PESO',
                'CONSUMO INTERNO', 'CAJA',
            }:
                shift.observaciones.append(ObservacionEntry(
                    flavor=name_obs_str,
                    nota=name_obs_str,
                ))
            continue

        # ── VENTAS SIN PESO (cols D-F) ────────────────────────────────────────
        # Detect "VENTA ANTES DEL PESAJE" / "VENTA DESPUES DEL PESO" header rows
        if col_d is not None and isinstance(col_d, str):
            kw = col_d.upper()
            if 'VENTA' in kw and ('PESAJE' in kw or 'PESO' in kw):
                continue  # skip header rows

        for cell_val in (col_d, col_e, col_f):
            entries = _parse_ventas_text(cell_val)
            shift.ventas_sin_peso.extend(entries)

        # ── CONSUMO INTERNO (cols G-I) ────────────────────────────────────────
        emp = _is_employee(col_g)
        if emp:
            desc_parts = []
            for v in (col_h, col_i):
                if v is not None and not isinstance(v, datetime.datetime):
                    desc_parts.append(str(v).strip())
            description = ' '.join(p for p in desc_parts if p)
            grams = text_to_grams(description)
            if description:
                shift.consumos.append(ConsumoEntry(
                    employee=str(col_g).strip(),
                    description=description,
                    grams=grams,
                ))
        # Also check if employee name is in col D (some sheets use wrong column)
        emp2 = _is_employee(col_d)
        if emp2 and not emp:
            desc_parts = []
            for v in (col_e, col_f, col_g):
                if v is not None and not isinstance(v, datetime.datetime):
                    s = str(v).strip()
                    if s:
                        desc_parts.append(s)
            description = ' '.join(desc_parts)
            grams = text_to_grams(description)
            if description:
                shift.consumos.append(ConsumoEntry(
                    employee=str(col_d).strip(),
                    description=description,
                    grams=grams,
                ))


# ══════════════════════════════════════════════════════════════════════════════
# V2 — Raw parsing with SlotValue preservation (additive, doesn't change V1)
# ══════════════════════════════════════════════════════════════════════════════

def _parse_raw_flavor(ws, row: int) -> RawFlavorObs:
    """Extract one flavor row preserving column positions as SlotValues.
    Reads ALL numeric values in columns D-K, tagging by column range."""
    name_raw = ws.cell(row, 1).value
    name = str(name_raw).strip() if name_raw else ''
    abierta = _safe_float(ws.cell(row, 2).value) or 0.0
    celiaca = _safe_float(ws.cell(row, 3).value) or 0.0

    slots = []
    # Columns D(4) through I(9) = cerrada slots
    for c in range(4, 10):
        v = _safe_float(ws.cell(row, c).value)
        if v is not None:
            col_letter = chr(64 + c)  # 4='D', 5='E', etc.
            slots.append(SlotValue(column=col_letter, value=v, slot_type='cerrada'))

    # Columns J(10) through K(11) = entrante slots
    for c in range(10, 12):
        v = _safe_float(ws.cell(row, c).value)
        if v is not None:
            col_letter = chr(64 + c)
            slots.append(SlotValue(column=col_letter, value=v, slot_type='entrante'))

    return RawFlavorObs(name=name, abierta=abierta, celiaca=celiaca, slots=slots)


def _vdp_cell_to_text(val) -> Optional[str]:
    """Convert a VDP cell value to text, handling Excel datetime misparsing.
    Excel interprets '1/4' as April 1st, '1/2' as Feb 1st or Jan 2nd, etc.
    We detect these and convert back to the fraction string."""
    if val is None:
        return None
    if isinstance(val, str):
        t = val.strip()
        return t if t else None
    if isinstance(val, datetime.datetime):
        # Excel misparses fractions as dates: "1/4" -> April 1, "1/2" -> Feb 1
        # Reconstruct: day/month (European format, which matches the locale)
        d, m = val.day, val.month
        # Common fractions: 1/2, 1/4 — always day=1, month=2 or 4
        if d == 1 and m in (2, 4):
            return f'1/{m}'
        # Less common but possible: 1/3, 3/4, etc.
        return f'{d}/{m}'
    # Numeric values in VDP area are pre-computed gram totals, not text
    # descriptions. Skip them — only strings and datetime misparsings are VDP entries.
    return None


def _parse_vdp_texts(ws, postres_row: int, max_row: int) -> list:
    """Extract raw VDP text strings from POSTRES section.
    Scans the full VDP rectangle (cols D-F) from POSTRES+1 until
    OBSERVACIONES or DELIVERY row. Handles Excel datetime misparsing
    of fractions like '1/4' → April 1st."""
    texts = []
    for r in range(postres_row + 1, max_row + 1):
        # Check col D first for section boundary markers
        col_d = ws.cell(r, 4).value
        if col_d is not None and isinstance(col_d, str):
            kw = col_d.upper()
            if 'OBSERVACION' in kw or 'DELIVERY' in kw:
                break
            if 'VENTA' in kw and ('PESAJE' in kw or 'PESO' in kw):
                # Skip the header row, but still check E/F below
                col_d = None  # don't add this cell

        # Collect from cols D, E, F
        for c_idx in [4, 5, 6]:
            val = ws.cell(r, c_idx).value if c_idx != 4 else col_d
            text = _vdp_cell_to_text(val)
            if text and text not in ('DIA', 'NOCHE'):
                texts.append(text)
    return texts


def load_shifts_v2(path: str):
    """Load all shifts as RawShift objects + separate annotations.
    Returns (list[RawShift], list[V2ShiftAnnotations]).
    STOCK sheets are included with is_stock_sheet=True but empty flavors."""
    import re as _re
    wb = openpyxl.load_workbook(path, data_only=True)
    shifts = []
    annotations = []

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]

        # Detect STOCK sheets
        name_upper = sheet_name.strip().upper()
        is_stock = name_upper.startswith('STOCK')

        if is_stock:
            shifts.append(RawShift(
                name=sheet_name.strip(),
                index=idx,
                is_stock_sheet=True,
            ))
            annotations.append(V2ShiftAnnotations(shift_name=sheet_name.strip()))
            continue

        # Only process sheets with A1='SABORES'
        if ws['A1'].value != 'SABORES':
            continue

        raw_shift = RawShift(name=sheet_name.strip(), index=idx)
        max_row = ws.max_row
        postres_row = None

        for r in range(2, max_row + 1):
            a_val = ws.cell(r, 1).value
            if a_val is not None and isinstance(a_val, str) and normalize_name(a_val).startswith('POSTRES'):
                postres_row = r
                break
            if a_val is None:
                b_val = _safe_float(ws.cell(r, 2).value)
                if b_val is not None and b_val > 10000:
                    continue
                continue
            if normalize_name(a_val) == 'SABORES':
                continue

            obs = _parse_raw_flavor(ws, r)
            name_norm = normalize_name(a_val)
            if name_norm and (obs.abierta > 0 or obs.celiaca > 0 or obs.slots):
                raw_shift.flavors[name_norm] = obs

        # Extract VDP texts
        if postres_row:
            raw_shift.vdp_texts = _parse_vdp_texts(ws, postres_row, max_row)

        # Extract annotations (separate concern)
        annot = V2ShiftAnnotations(shift_name=sheet_name.strip())
        # TODO: extract latas_cambiadas, consumos, observaciones as raw strings
        annotations.append(annot)

        shifts.append(raw_shift)

    return shifts, annotations


